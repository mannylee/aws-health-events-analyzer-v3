import boto3
import json
import re
import csv
import ast
from datetime import datetime, timedelta, timezone
from datetime import datetime, date
from io import BytesIO
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from botocore.exceptions import ClientError
import base64
import traceback
import os
import uuid

# Environment variables for configuration
BEDROCK_MODEL_ID = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-3-5-sonnet-20240620-v1:0')
BEDROCK_TEMPERATURE = float(os.environ.get('BEDROCK_TEMPERATURE', '0.1'))
BEDROCK_TOP_P = float(os.environ.get('BEDROCK_TOP_P', '0.9'))
BEDROCK_MAX_TOKENS = int(os.environ.get('BEDROCK_MAX_TOKENS', '4000'))
EXCEL_FILENAME_TEMPLATE = os.environ.get('EXCEL_FILENAME_TEMPLATE', 'AWS_Health_Events_Analysis_{date}_{time}.xlsx')
customer_name = os.environ.get('CUSTOMER_NAME', 'Notification')
excluded_services_str = os.environ.get('EXCLUDED_SERVICES', '')
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME', '')
S3_KEY_PREFIX = os.environ.get('S3_KEY_PREFIX', '')
REPORTS_BUCKET = os.environ.get('REPORTS_BUCKET', '')
USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING = os.environ.get('USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING', 'false').lower() == 'true'
USE_CUSTOM_ACCOUNT_EMAIL_MAPPING = os.environ.get('USE_CUSTOM_ACCOUNT_EMAIL_MAPPING', 'false').lower() == 'true'
OVERRIDE_S3_HEALTH_EVENTS_ARN = os.environ.get('OVERRIDE_S3_HEALTH_EVENTS_ARN', '')

excluded_services = [s.strip() for s in excluded_services_str.split(',') if s.strip()]


def is_ses_in_sandbox_mode():
    """
    Check if SES is running in sandbox mode by examining sending quota.
    
    Returns:
        bool: True if SES is in sandbox mode, False if in production mode
    """
    try:
        ses_client = boto3.client('ses')
        quota = ses_client.get_send_quota()
        
        # Sandbox mode indicators:
        # - Max24HourSend: 200 emails per 24 hours
        # - MaxSendRate: 1 email per second
        max_24_hour_send = quota.get('Max24HourSend', 0)
        max_send_rate = quota.get('MaxSendRate', 0)
        
        is_sandbox = (max_24_hour_send <= 200 and max_send_rate <= 1)
        
        # print(f"SES Status - Max24HourSend: {max_24_hour_send}, MaxSendRate: {max_send_rate}")
        print(f"SES is in {'SANDBOX' if is_sandbox else 'PRODUCTION'} mode")
        
        return is_sandbox
        
    except Exception as e:
        print(f"Warning: Could not determine SES mode, assuming sandbox: {e}")
        return True  # Default to sandbox mode for safety


def should_verify_recipient_email():
    """
    Determine if recipient email verification is required based on SES mode.
    
    Returns:
        bool: True if verification is required (sandbox mode), False otherwise
    """
    return is_ses_in_sandbox_mode()


def get_custom_account_email_mapping_from_dynamodb():
    """
    Fetch custom account-email mapping from DynamoDB table
    
    Returns:
        tuple: (account_to_email_mapping, email_to_accounts_mapping)
            - account_to_email_mapping: dict mapping account IDs to email addresses
            - email_to_accounts_mapping: dict mapping email addresses to lists of account IDs
    """
    # Check if custom mapping is enabled
    if not USE_CUSTOM_ACCOUNT_EMAIL_MAPPING:
        print("Custom account email mapping is disabled")
        return {}, {}
    
    table_name = os.environ.get('ACCOUNT_EMAIL_MAPPING_TABLE', '')
    if not table_name:
        print("Custom mapping is enabled but no DynamoDB table name configured")
        return {}, {}
    
    try:
        print(f"Fetching custom account email mapping from DynamoDB table: {table_name}")
        dynamodb = boto3.resource('dynamodb')
        table = dynamodb.Table(table_name)
        
        # Scan the table to get all mappings
        response = table.scan()
        items = response['Items']
        
        # Handle pagination if there are many items
        while 'LastEvaluatedKey' in response:
            print("Fetching additional items from DynamoDB (pagination)...")
            response = table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
            items.extend(response['Items'])
        
        print(f"Successfully loaded {len(items)} account mappings from DynamoDB")
        
        # Convert to both formats for compatibility with existing code
        account_to_email = {}
        email_to_accounts = defaultdict(list)
        
        for item in items:
            account_id = item.get('AccountId', '').strip()
            email = item.get('Email', '').strip()
            team_name = item.get('TeamName', '')
            environment = item.get('Environment', '')
            
            if not account_id or not email:
                print(f"Warning: Skipping invalid mapping - AccountId: '{account_id}', Email: '{email}'")
                continue
            
            account_to_email[account_id] = email
            email_to_accounts[email].append(account_id)
            
            # Log the mapping with additional context
            context_info = []
            if team_name:
                context_info.append(f"Team: {team_name}")
            if environment:
                context_info.append(f"Env: {environment}")
            
            context_str = f" ({', '.join(context_info)})" if context_info else ""
            print(f"Custom mapping: {account_id} -> {email}{context_str}")
        
        print(f"Created account-to-email mapping for {len(account_to_email)} accounts")
        print(f"Found {len(email_to_accounts)} unique email addresses")
        
        # Log email-to-accounts mapping for verification
        for email, accounts in email_to_accounts.items():
            print(f"Email {email} will receive events for accounts: {', '.join(accounts)}")
        
        return account_to_email, dict(email_to_accounts)
        
    except ClientError as e:
        error_code = e.response['Error']['Code']
        if error_code == 'ResourceNotFoundException':
            print(f"DynamoDB table {table_name} not found")
        elif error_code == 'AccessDeniedException':
            print(f"Access denied when trying to read DynamoDB table {table_name}. Ensure Lambda has dynamodb:Scan permission.")
        else:
            print(f"Error accessing DynamoDB table: {error_code} - {e.response['Error']['Message']}")
        return {}, {}
    except Exception as e:
        print(f"Error fetching custom mapping from DynamoDB: {str(e)}")
        traceback.print_exc()
        return {}, {}


def get_organization_account_email_mapping():
    """
    Get account-to-email mapping from AWS Organizations by looking up account contact information.
    
    Returns:
        tuple: (account_to_email_mapping, email_to_accounts_mapping)
            - account_to_email_mapping: dict mapping account IDs to email addresses
            - email_to_accounts_mapping: dict mapping email addresses to lists of account IDs
    """
    account_to_email = {}
    email_to_accounts = defaultdict(list)
    
    if not USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING:
        print("Organization account email mapping is disabled")
        return account_to_email, dict(email_to_accounts)
    
    try:
        print("Fetching account email mappings from AWS Organizations...")
        
        # Create Organizations client
        org_client = boto3.client('organizations')
        
        # List all accounts in the organization
        paginator = org_client.get_paginator('list_accounts')
        
        for page in paginator.paginate():
            for account in page['Accounts']:
                account_id = account['Id']
                account_name = account.get('Name', 'Unknown')
                account_email = account.get('Email')
                
                if account_email:
                    account_to_email[account_id] = account_email
                    email_to_accounts[account_email].append(account_id)
                    print(f"Account mapping: {account_id} ({account_name}) -> {account_email}")
                else:
                    print(f"Warning: No email found for account {account_id} ({account_name})")
        
        print(f"Retrieved {len(account_to_email)} account email mappings from Organizations")
        print(f"Found {len(email_to_accounts)} unique email addresses")
        
        for email, accounts in email_to_accounts.items():
            account_names = []
            for acc_id in accounts:
                # Find account name for better logging
                try:
                    account_info = org_client.describe_account(AccountId=acc_id)
                    account_names.append(f"{acc_id} ({account_info['Account'].get('Name', 'Unknown')})")
                except:
                    account_names.append(acc_id)
            print(f"Email {email} will receive events for accounts: {', '.join(account_names)}")
        
    except ClientError as e:
        error_code = e.response['Error']['Code']
        if error_code == 'AccessDeniedException':
            print("Error: Access denied when trying to access AWS Organizations. Ensure the Lambda has organizations:ListAccounts and organizations:DescribeAccount permissions.")
        elif error_code == 'AWSOrganizationsNotInUseException':
            print("Error: AWS Organizations is not enabled for this account.")
        else:
            print(f"Error accessing AWS Organizations: {error_code} - {e.response['Error']['Message']}")
    except Exception as e:
        print(f"Error fetching account mappings from Organizations: {str(e)}")
        traceback.print_exc()
    
    return account_to_email, dict(email_to_accounts)


def get_combined_account_email_mapping():
    """
    Get account-email mapping using hybrid approach:
    1. Start with AWS Organizations mapping (if enabled)
    2. Override with custom DynamoDB mapping (if enabled)
    3. DynamoDB takes precedence for accounts that exist in both sources
    
    Returns:
        tuple: (account_to_email_mapping, email_to_accounts_mapping)
            - account_to_email_mapping: dict mapping account IDs to email addresses
            - email_to_accounts_mapping: dict mapping email addresses to lists of account IDs
    """
    print(f"Account email mapping configuration:")
    print(f"  - Custom DynamoDB mapping: {'ENABLED' if USE_CUSTOM_ACCOUNT_EMAIL_MAPPING else 'DISABLED'}")
    print(f"  - Organizations mapping: {'ENABLED' if USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING else 'DISABLED'}")
    
    # Initialize combined mappings
    combined_account_to_email = {}
    combined_email_to_accounts = defaultdict(list)
    
    # Step 1: Get Organizations mapping as base (if enabled)
    if USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING:
        print("Fetching AWS Organizations mapping as base...")
        org_account_to_email, org_email_to_accounts = get_organization_account_email_mapping()
        
        if org_account_to_email:
            combined_account_to_email.update(org_account_to_email)
            for email, accounts in org_email_to_accounts.items():
                combined_email_to_accounts[email].extend(accounts)
            print(f"Added {len(org_account_to_email)} accounts from Organizations mapping")
        else:
            print("Organizations mapping is enabled but no mappings found")
    
    # Step 2: Override with DynamoDB mapping (if enabled)
    if USE_CUSTOM_ACCOUNT_EMAIL_MAPPING:
        print("Fetching custom DynamoDB mapping for overrides...")
        custom_account_to_email, custom_email_to_accounts = get_custom_account_email_mapping_from_dynamodb()
        
        if custom_account_to_email:
            # Track overrides for logging
            overridden_accounts = []
            new_accounts = []
            
            for account_id, new_email in custom_account_to_email.items():
                old_email = combined_account_to_email.get(account_id)
                
                if old_email and old_email != new_email:
                    # This account exists in Organizations but with different email
                    overridden_accounts.append(f"{account_id}: {old_email} -> {new_email}")
                    # Remove from old email list
                    if old_email in combined_email_to_accounts:
                        combined_email_to_accounts[old_email] = [
                            acc for acc in combined_email_to_accounts[old_email] if acc != account_id
                        ]
                        # Clean up empty email lists
                        if not combined_email_to_accounts[old_email]:
                            del combined_email_to_accounts[old_email]
                elif not old_email:
                    # This is a new account not in Organizations
                    new_accounts.append(f"{account_id} -> {new_email}")
                
                # Apply the DynamoDB mapping
                combined_account_to_email[account_id] = new_email
                if account_id not in combined_email_to_accounts[new_email]:
                    combined_email_to_accounts[new_email].append(account_id)
            
            print(f"Added {len(custom_account_to_email)} accounts from DynamoDB mapping")
            
            if overridden_accounts:
                print(f"DynamoDB overrode {len(overridden_accounts)} Organizations mappings:")
                for override in overridden_accounts:
                    print(f"  Override: {override}")
            
            if new_accounts:
                print(f"DynamoDB added {len(new_accounts)} new accounts not in Organizations:")
                for new_account in new_accounts:
                    print(f"  New: {new_account}")
        else:
            print("Custom DynamoDB mapping is enabled but no mappings found")
    
    # Convert defaultdict back to regular dict
    combined_email_to_accounts = dict(combined_email_to_accounts)
    
    # Summary logging
    if combined_account_to_email:
        print(f"Final combined mapping: {len(combined_account_to_email)} accounts across {len(combined_email_to_accounts)} email addresses")
        
        # Log final email-to-accounts mapping for verification
        for email, accounts in combined_email_to_accounts.items():
            print(f"Email {email} will receive events for accounts: {', '.join(accounts)}")
        
        return combined_account_to_email, combined_email_to_accounts
    else:
        print("No account-email mapping configured - all events will go to default recipients")
        return {}, {}


def expand_events_by_account(events):
    """
    Expands events that affect multiple accounts into separate event records for each account.
    Fetches affected accounts if not already specified.
    
    Args:
        events (list): List of event dictionaries
        
    Returns:
        list: Expanded list of event dictionaries
    """
    expanded_events = []
    
    # Health API should go to us-east-1
    health_client = boto3.client('health', region_name='us-east-1')
    
    for event in events:
        # Get the account ID string which may contain multiple comma-separated IDs
        account_id_str = event.get('accountId', '')
        event_arn = event.get('arn', '')
        
        # If no account ID or it's N/A, try to fetch affected accounts
        if not account_id_str or account_id_str == 'N/A':
            try:
                print(f"Fetching affected accounts for event: {event.get('eventTypeCode', 'unknown')}")
                response = health_client.describe_affected_accounts_for_organization(
                    eventArn=event_arn
                )
                affected_accounts = response.get('affectedAccounts', [])
                
                if affected_accounts:
                    # If multiple accounts are affected, join them with commas
                    account_id_str = ', '.join(affected_accounts)
                    event['accountId'] = account_id_str  # Update the event with the account IDs
                    print(f"Found affected accounts: {account_id_str}")
                else:
                    print("No affected accounts found")
                    expanded_events.append(event)  # Keep the event as is
                    continue
            except Exception as e:
                print(f"Error fetching affected accounts: {str(e)}")
                expanded_events.append(event)  # Keep the event as is
                continue
        
        # If no comma in the string, it's a single account or none
        if ',' not in account_id_str:
            expanded_events.append(event)
            continue
        
        # Split the account IDs and create a separate event for each
        account_ids = [aid.strip() for aid in account_id_str.split(',')]
        print(f"Expanding event {event_arn} for {len(account_ids)} accounts: {account_ids}")
        
        for account_id in account_ids:
            # Create a copy of the event for this specific account
            account_event = event.copy()
            account_event['accountId'] = account_id
            expanded_events.append(account_event)
    
    print(f"Expanded {len(events)} events to {len(expanded_events)} account-specific events")
    return expanded_events


def download_s3_file(s3_location):
    """
    Download file from S3 location (supports both CSV and Excel files)
    
    Args:
        s3_location (str): S3 location in format s3://bucket/key or bucket/key
        
    Returns:
        BytesIO: File content as BytesIO object
    """
    try:
        # Parse S3 location
        if s3_location.startswith('s3://'):
            s3_location = s3_location[5:]  # Remove s3:// prefix
        
        parts = s3_location.split('/', 1)
        if len(parts) != 2:
            raise ValueError(f"Invalid S3 location format: {s3_location}")
        
        bucket_name, key = parts
        
        print(f"Downloading file from S3: bucket={bucket_name}, key={key}")
        
        # Create S3 client
        s3_client = boto3.client('s3')
        
        # Download file to BytesIO
        file_buffer = BytesIO()
        s3_client.download_fileobj(bucket_name, key, file_buffer)
        file_buffer.seek(0)
        
        print(f"Successfully downloaded file from S3")
        return file_buffer, key
        
    except Exception as e:
        print(f"Error downloading file from S3: {str(e)}")
        raise


def parse_health_events_file(file_buffer, filename):
    """
    Parse CSV or Excel file containing health events and convert to Health API format
    
    Args:
        file_buffer (BytesIO): File content
        filename (str): Name of the file to determine format
        
    Returns:
        list: List of health events in Health API format
    """
    try:
        # Determine file type from extension
        is_csv = filename.lower().endswith('.csv')
        
        if is_csv:
            print("Parsing CSV file for health events...")
            return parse_csv_health_events(file_buffer)
        else:
            print("Parsing Excel file for health events...")
            return parse_excel_health_events(file_buffer)
        
    except Exception as e:
        print(f"Error parsing health events file: {str(e)}")
        traceback.print_exc()
        raise


def _get_column_mapping():
    """
    Get the standard column mapping for health events
    
    Returns:
        dict: Mapping from file column names to Health API field names
    """
    return {
        'domain': 'domain',  # Keep domain for reference
        'account_id': 'accountId',
        'event_arn': 'arn',
        'event_type_code': 'eventTypeCode',
        'event_type_category': 'eventTypeCategory',
        'service': 'service',
        'region': 'region',
        'availability_zone': 'availabilityZone',
        'start_time': 'startTime',
        'end_time': 'endTime',
        'last_updated_time': 'lastUpdatedTime',
        'status_code': 'statusCode',
        'event_scope_code': 'eventScopeCode',
        'affected_entities_count': 'affectedEntitiesCount',
        'affected_entities': 'affectedEntities',
        'details': 'details'  # Special handling - extracts description only
    }


def _read_csv_data(csv_buffer):
    """
    Read CSV data and return headers and rows
    
    Args:
        csv_buffer (BytesIO): CSV file content
        
    Returns:
        tuple: (headers, rows_iterator)
    """
    import csv
    
    csv_buffer.seek(0)
    csv_content = csv_buffer.read().decode('utf-8-sig')  # Handle BOM if present
    csv_reader = csv.DictReader(csv_content.splitlines())
    
    headers = list(csv_reader.fieldnames)
    rows = [(row_num, row) for row_num, row in enumerate(csv_reader, start=2)]
    
    return headers, rows


def _read_excel_data(excel_buffer):
    """
    Read Excel data and return headers and rows
    
    Args:
        excel_buffer (BytesIO): Excel file content
        
    Returns:
        tuple: (headers, rows_iterator)
    """
    # Load the workbook
    workbook = load_workbook(excel_buffer, data_only=True)
    
    # Try to find the worksheet with health events data
    sheet_names_to_try = ['Sheet1', 'Events', 'Health Events', 'AWS Health Events']
    worksheet = None
    
    for sheet_name in sheet_names_to_try:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            break
    
    if worksheet is None:
        worksheet = workbook.active
        print(f"Using worksheet: {worksheet.title}")
    else:
        print(f"Found worksheet: {worksheet.title}")
    
    # Read the header row
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append('')
    
    # Convert Excel rows to dictionary format similar to CSV
    rows = []
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        # Create dictionary mapping headers to values
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row) and row[i] is not None:
                row_dict[header] = row[i]
        
        rows.append((row_num, row_dict))
    
    return headers, rows


def _process_health_event_row(row_data, row_num, column_mapping, file_type):
    """
    Process a single row of health event data into Health API format
    
    Args:
        row_data (dict): Row data as dictionary
        row_num (int): Row number for error reporting
        column_mapping (dict): Column mapping dictionary
        file_type (str): 'csv' or 'excel' for error messages
        
    Returns:
        dict: Processed health event
    """
    import ast
    
    event = {}
    
    # Extract fields based on mapping
    for file_col, api_field in column_mapping.items():
        if file_col in row_data and row_data[file_col] is not None:
            value = str(row_data[file_col]).strip()
            if not value:  # Skip empty values
                continue
            
            # Handle special fields that need parsing
            if file_col == 'details' and value:
                try:
                    # Parse the details field which contains a dictionary as string
                    details_dict = ast.literal_eval(value)
                    if isinstance(details_dict, dict):
                        # Extract description from details - this is the main field we need
                        if 'description' in details_dict:
                            event['eventDescription'] = details_dict['description']
                        # Also extract affected_entities if present in details and not already set
                        if 'affected_entities' in details_dict and 'affectedEntities' not in event:
                            event['affectedEntities'] = details_dict['affected_entities']
                        # Don't store the full details blob to keep data size manageable
                except (ValueError, SyntaxError) as e:
                    print(f"Warning: Could not parse details field for row {row_num}: {e}")
                    # Try to extract at least the description if it's a simple string
                    if 'description' in value.lower():
                        event['eventDescription'] = value[:1000] + "..." if len(value) > 1000 else value
                    else:
                        event['eventDescription'] = value[:500] + "..." if len(value) > 500 else value
            elif file_col == 'affected_entities' and value:
                try:
                    # This might be a string representation of a list or just a string
                    if value.startswith('[') or value.startswith('arn:'):
                        event[api_field] = value
                    else:
                        event[api_field] = value
                except Exception as e:
                    print(f"Warning: Could not parse affected_entities for row {row_num}: {e}")
                    event[api_field] = value
            elif file_col == 'affected_entities_count' and value:
                try:
                    # Convert to integer
                    event[api_field] = int(value)
                except ValueError:
                    print(f"Warning: Could not convert affected_entities_count to int for row {row_num}: {value}")
                    event[api_field] = value
            else:
                event[api_field] = value
    
    # Set default values for required fields if missing
    if 'arn' not in event:
        event['arn'] = f"arn:aws:health:global::event/{file_type}-event-{row_num}"
    
    if 'eventTypeCode' not in event:
        event['eventTypeCode'] = 'UNKNOWN_EVENT_TYPE'
    
    if 'service' not in event:
        event['service'] = 'unknown'
    
    if 'region' not in event:
        event['region'] = 'global'
    
    if 'statusCode' not in event:
        event['statusCode'] = 'closed'
    
    if 'eventTypeCategory' not in event:
        event['eventTypeCategory'] = 'accountNotification'
    
    # Handle datetime fields
    current_time = datetime.now(timezone.utc).isoformat()
    if 'startTime' not in event:
        event['startTime'] = current_time
    
    if 'lastUpdatedTime' not in event:
        event['lastUpdatedTime'] = current_time
    
    # Ensure accountId is set
    if 'accountId' not in event:
        event['accountId'] = '123456789012'  # Default test account
    
    return event


def parse_csv_health_events(csv_buffer):
    """
    Parse CSV file containing health events and convert to Health API format
    
    Args:
        csv_buffer (BytesIO): CSV file content
        
    Returns:
        list: List of health events in Health API format
    """
    try:
        print("Parsing CSV file for health events...")
        
        # Read CSV data
        headers, rows = _read_csv_data(csv_buffer)
        column_mapping = _get_column_mapping()
        
        print(f"CSV headers found: {headers}")
        
        # Validate that we have the expected columns
        expected_columns = ['account_id', 'event_arn', 'event_type_code', 'service']
        missing_columns = [col for col in expected_columns if col not in headers]
        if missing_columns:
            print(f"Warning: Missing expected columns: {missing_columns}")
        
        # Process each row
        events = []
        for row_num, row_data in rows:
            if not any(row_data.values()):  # Skip empty rows
                continue
            
            event = _process_health_event_row(row_data, row_num, column_mapping, 'csv')
            events.append(event)
        
        print(f"Parsed {len(events)} events from CSV file")
        
        # Log first event for debugging
        if events:
            print(f"Sample event keys: {list(events[0].keys())}")
            print(f"Sample event: {json.dumps(events[0], indent=2, default=str)}")
        
        return events
        
    except Exception as e:
        print(f"Error parsing CSV file: {str(e)}")
        traceback.print_exc()
        raise


def parse_excel_health_events(excel_buffer):
    """
    Parse Excel file containing health events and convert to Health API format
    
    Args:
        excel_buffer (BytesIO): Excel file content
        
    Returns:
        list: List of health events in Health API format
    """
    try:
        print("Parsing Excel file for health events...")
        
        # Read Excel data
        headers, rows = _read_excel_data(excel_buffer)
        column_mapping = _get_column_mapping()
        
        print(f"Excel headers found: {headers}")
        
        # Validate that we have the expected columns
        expected_columns = ['account_id', 'event_arn', 'event_type_code', 'service']
        missing_columns = [col for col in expected_columns if col not in headers]
        if missing_columns:
            print(f"Warning: Missing expected columns: {missing_columns}")
        
        # Process each row
        events = []
        for row_num, row_data in rows:
            if not any(row_data.values()):  # Skip empty rows
                continue
            
            event = _process_health_event_row(row_data, row_num, column_mapping, 'excel')
            events.append(event)
        
        print(f"Parsed {len(events)} events from Excel file")
        
        # Log first event for debugging
        if events:
            print(f"Sample event keys: {list(events[0].keys())}")
            print(f"Sample event: {json.dumps(events[0], indent=2, default=str)}")
        
        return events
        
    except Exception as e:
        print(f"Error parsing Excel file: {str(e)}")
        traceback.print_exc()
        raise


def lambda_handler(event, context):
    print("Starting execution...")
    
    try:
        # Get all configuration from environment variables
        analysis_window_days = int(os.environ['ANALYSIS_WINDOW_DAYS'])
        
        # Get event categories to process from environment variable
        event_categories_to_process = []
        if 'EVENT_CATEGORIES' in os.environ and os.environ['EVENT_CATEGORIES'].strip():
            event_categories_to_process = [cat.strip() for cat in os.environ['EVENT_CATEGORIES'].split(',')]
            print(f"Will only process these event categories: {event_categories_to_process}")
        else:
            print("No EVENT_CATEGORIES specified, will process all event categories")

        excluded_services_str = os.environ.get('EXCLUDED_SERVICES', '')
        excluded_services = [s.strip() for s in excluded_services_str.split(',') if s.strip()]
        
        if excluded_services:
            print(f"Excluding services from analysis: {excluded_services}")
        
        # Set up time range for filtering using environment variable
        bedrock_client = get_bedrock_client()
        start_time = datetime.now(timezone.utc) - timedelta(days=analysis_window_days)
        # No end_time cap - look into the future indefinitely
        
        print(f"Fetching events from {start_time} onwards (no end date)")
        
        # Check if we should use S3 health events instead of Health API
        if OVERRIDE_S3_HEALTH_EVENTS_ARN:
            print(f"OVERRIDE_S3_HEALTH_EVENTS_ARN is set to: {OVERRIDE_S3_HEALTH_EVENTS_ARN}")
            print("Using S3 file for health events instead of Health API")
            
            try:
                # Download and parse file from S3 (CSV or Excel)
                file_buffer, filename = download_s3_file(OVERRIDE_S3_HEALTH_EVENTS_ARN)
                all_events = parse_health_events_file(file_buffer, filename)
                filtered_count = 0
                
                print(f"Successfully loaded {len(all_events)} events from S3 file")
                
            except Exception as e:
                print(f"Error processing S3 file: {str(e)}")
                traceback.print_exc()
                return {
                    'statusCode': 500,
                    'body': json.dumps({'error': f'Failed to process S3 file: {str(e)}'})
                }
        else:
            print("Using AWS Health API for events")
            
            # Format dates properly for the API
            formatted_start = start_time.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
            # No end time formatting needed - looking into future indefinitely
            
            # Initialize AWS Health client
            health_client = boto3.client('health', region_name='us-east-1')
            
            # Initialize variables for event collection
            all_events = []
            filtered_count = 0
            
            try:
                # Check if we should use organization view or account view
                use_org_view = is_org_view_enabled()
                
                if use_org_view:
                    print("Using AWS Health Organization View")
                
                    # CHANGE 1: Fetch non-closed events (open and upcoming only)
                    non_closed_filter = {
                        'startTime': {'from': formatted_start},
                        'eventStatusCodes': ['open', 'upcoming']  # Only non-closed events
                    }
                    
                    # Add event type categories filter if specified
                    if event_categories_to_process:
                        non_closed_filter['eventTypeCategories'] = event_categories_to_process
                    
                    print(f"Fetching NON-CLOSED events (open/upcoming) with filter: {non_closed_filter}")
                    response = health_client.describe_events_for_organization(
                        filter=non_closed_filter,
                        maxResults=100
                    )
                    
                    if 'events' in response:
                        all_events.extend(response['events'])
                        print(f"Retrieved {len(response.get('events', []))} non-closed events")
                    
                    # Handle pagination for non-closed events
                    while 'nextToken' in response and response['nextToken']:
                        print(f"Found nextToken for non-closed events, fetching more...")
                        if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                            print("Approaching Lambda timeout, stopping pagination")
                            break
                            
                        response = health_client.describe_events_for_organization(
                            filter=non_closed_filter,
                            maxResults=100,
                            nextToken=response['nextToken']
                        )
                        
                        if 'events' in response:
                            all_events.extend(response['events'])
                            print(f"Retrieved {len(response.get('events', []))} additional non-closed events")
                
                else:
                    print("Using AWS Health Account View")
                    
                    # CHANGE 3: Same approach for account view - fetch non-closed events only
                    non_closed_filter = {
                        'startTime': {'from': formatted_start},
                        'eventStatusCodes': ['open', 'upcoming']  # Only non-closed events
                    }
                    
                    # Add event type categories filter if specified
                    if event_categories_to_process:
                        non_closed_filter['eventTypeCategories'] = event_categories_to_process
                    
                    print(f"Fetching NON-CLOSED events (open/upcoming) with filter: {non_closed_filter}")
                    response = health_client.describe_events(
                        filter=non_closed_filter,
                        maxResults=100
                    )
                    
                    if 'events' in response:
                        all_events.extend(response['events'])
                    
                    # Handle pagination for non-closed events
                    while 'nextToken' in response and response['nextToken']:
                        if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                            print("Approaching Lambda timeout, stopping pagination")
                            break
                            
                        response = health_client.describe_events(
                            filter=non_closed_filter,
                            maxResults=100,
                            nextToken=response['nextToken']
                        )
                        
                        if 'events' in response:
                            all_events.extend(response['events'])
        
            except ClientError as e:
                if e.response['Error']['Code'] == 'SubscriptionRequiredException':
                    print("Health Organization View is not enabled. Falling back to account-specific view.")
                    
                    # CHANGE 5: Same approach for fallback - fetch closed events
                    closed_filter = {
                        'startTime': {'from': formatted_start},
                        # No endTime - look into the future indefinitely
                        'eventStatusCodes': ['closed', 'upcoming']
                    }
                    
                    # Add event type categories filter if specified
                    if event_categories_to_process:
                        closed_filter['eventTypeCategories'] = event_categories_to_process
                    
                    print(f"Fetching CLOSED events with filter: {closed_filter}")
                    closed_response = health_client.describe_events(
                        filter=closed_filter,
                        maxResults=100
                    )
                    
                    if 'events' in closed_response:
                        all_events.extend(closed_response['events'])
                    
                    # Handle pagination for closed events
                    while 'nextToken' in closed_response and closed_response['nextToken']:
                        if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                            print("Approaching Lambda timeout, stopping pagination")
                            break
                            
                        closed_response = health_client.describe_events(
                            filter=closed_filter,
                            maxResults=100,
                            nextToken=closed_response['nextToken']
                        )
                        
                        if 'events' in closed_response:
                            all_events.extend(closed_response['events'])
                    
                    # CHANGE 6: Fetch open events with only start date filter
                    open_filter = {
                        'startTime': {'from': formatted_start},  # Started on or after start date
                        'eventStatusCodes': ['open']  # Only open events
                    }
                    
                    # Add event type categories filter if specified
                    if event_categories_to_process:
                        open_filter['eventTypeCategories'] = event_categories_to_process
                    
                    print(f"Fetching OPEN events with filter: {open_filter}")
                    open_response = health_client.describe_events(
                        filter=open_filter,
                        maxResults=100
                    )
                    
                    if 'events' in open_response:
                        all_events.extend(open_response['events'])
                    
                    # Handle pagination for open events
                    while 'nextToken' in open_response and open_response['nextToken']:
                        if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                            print("Approaching Lambda timeout, stopping pagination")
                            break
                            
                        open_response = health_client.describe_events(
                            filter=open_filter,
                            maxResults=100,
                            nextToken=open_response['nextToken']
                        )
                        
                        if 'events' in open_response:
                            all_events.extend(open_response['events'])
                else:
                    raise
        
        # CHANGE 7: Remove duplicates by ARN
        unique_events = {}
        for item in all_events:
            arn = item.get('arn')
            if arn and arn not in unique_events:
                unique_events[arn] = item
        
        all_events = list(unique_events.values())

        # Filter out excluded services post-retrieval
        if excluded_services:
            filtered_events = [e for e in all_events if e.get('service') not in excluded_services]
            print(f"Filtered out {len(all_events) - len(filtered_events)} events from excluded services")
            all_events = filtered_events
        
        # NEW STEP: Expand events for multiple accounts
        all_events_original = all_events.copy()
        all_events_expanded = expand_events_by_account(all_events)
        print(f"Expanded {len(all_events_original)} events to {len(all_events_expanded)} account-specific events")
        
        items_count = len(all_events_original)  # Keep the original count for reporting
        print(f"Fetched {items_count} unique events from AWS Health API (expanded to {len(all_events_expanded)} account-specific events)")
        
        if len(all_events_expanded) == 0:
            print("No events found with filter")
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'No events matched the filter criteria',
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
        
        # Process events directly from the API results
        events_analysis = []
        event_categories = defaultdict(int)
        raw_events = []  # Store raw event data for Excel
        
        # Process each event from the expanded API results
        for item in all_events_expanded:
            if context.get_remaining_time_in_millis() > 10000:
                # Check if we should process this event category
                event_type_category = item.get('eventTypeCategory', '')
                
                # Skip events that don't match our configured categories (this is redundant since we're filtering in the API call,
                # but keeping it for consistency with the original code)
                if event_categories_to_process and event_type_category not in event_categories_to_process:
                    print(f"Skipping event {item.get('eventTypeCode', 'unknown')} with category {event_type_category} (not in configured categories)")
                    filtered_count += 1
                    continue
                
                print(f"Processing event: {item.get('eventTypeCode', 'unknown')} with category {event_type_category}")
                
                try:
                    # Store raw event data for Excel
                    raw_events.append(item)
                    
                    # Ensure we have the event ARN and standardize field name
                    event_arn = item.get('arn', '')
                    if event_arn:
                        item['eventArn'] = event_arn  # Standardize field name
                
                    # Extract account ID from ARN - this is already handled by the expansion function
                    account_id = item.get('accountId', 'N/A')
                    print(f"Processing with account ID: {account_id}")
                    
                    # Fetch additional details from Health API - now with a single account ID
                    health_data = fetch_health_event_details1(item.get('arn', ''), account_id)
                    
                    # Extract the actual description for analysis - IMPROVED EXTRACTION
                    actual_description = health_data['details'].get('eventDescription', {}).get('latestDescription', '')
                    
                    # If no description from Health API, try other possible fields
                    if not actual_description:
                        actual_description = (
                            item.get('eventDescription', '') or 
                            item.get('description', '') or 
                            item.get('message', '') or
                            'No description available'
                        )
                    
                    # Log the description we found
                    print(f"Using description (length: {len(actual_description)}): {actual_description[:100]}...")
                    
                    # Update the item with the actual description to improve analysis
                    item_with_description = item.copy()
                    item_with_description['description'] = actual_description
                    
                    analysis = analyze_event_with_bedrock(bedrock_client, item_with_description)
                    
                    categories = categorize_analysis(analysis)
                    if categories.get('critical', False):
                        event_categories['critical'] += 1
                    
                    risk_level = categories.get('risk_level', 'low')
                    event_categories[f"{risk_level}_risk"] += 1
                    
                    account_impact = categories.get('account_impact', 'low')
                    event_categories[f"{account_impact}_impact"] += 1
                    
                    # Create structured event data with both raw data and analysis
                    event_entry = {
                        "arn": item.get('arn', 'N/A'),
                        "eventArn": item.get('eventArn', item.get('arn', 'N/A')),  # Ensure eventArn is included
                        "event_type": item.get('eventTypeCode', 'N/A'),
                        "description": actual_description,
                        "region": item.get('region', 'N/A'),
                        "start_time": format_time(item.get('startTime', 'N/A')),
                        "last_update_time": format_time(item.get('lastUpdatedTime', 'N/A')),
                        "event_type_category": item.get('eventTypeCategory', 'N/A'),
                        "analysis_text": analysis,
                        "critical": categories.get('critical', False),
                        "risk_level": categories.get('risk_level', 'low'),
                        "accountId": account_id,  # Use the single account ID from the expanded event
                        "impact_analysis": categories.get('impact_analysis', ''),
                        "required_actions": categories.get('required_actions', ''),
                        "time_sensitivity": categories.get('time_sensitivity', 'Routine'),
                        "risk_category": categories.get('risk_category', 'Unknown'),
                        "consequences_if_ignored": categories.get('consequences_if_ignored', ''),
                        "affected_resources": extract_affected_resources(health_data['entities']),
                        "key_date": categories.get('key_date', 'N/A')
                    }
                    
                    events_analysis.append(event_entry)
                    print(f"Successfully analyzed event {len(events_analysis)}")
                except Exception as e:
                    print(f"Error analyzing event: {str(e)}")
                    traceback.print_exc()
            else:
                print("Approaching Lambda timeout, stopping event processing")
                break
        
        if events_analysis:
            print(f"Successfully analyzed {len(events_analysis)} events (filtered out {filtered_count} events)")
            
            # Create Excel report with structured data
            excel_buffer = create_excel_report_improved(events_analysis)
            
            # Generate summary HTML with filtering info
            summary_html = generate_summary_html(
                items_count,  # Use original count before expansion
                event_categories, 
                filtered_count, 
                event_categories_to_process if event_categories_to_process else None,
                events_analysis  
            )
            
            # Send emails with account-specific logic
            send_account_specific_emails(events_analysis, items_count, event_categories, filtered_count, event_categories_to_process)
            
            try:
                # Add CloudWatch metrics with error handling
                add_cloudwatch_metrics(event_categories, len(events_analysis), items_count, filtered_count)
            except Exception as e:
                print(f"Error publishing CloudWatch metrics: {str(e)}")
                print("Continuing execution despite metrics error")
            
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'total_events': items_count,  # Original event count
                    'total_expanded_events': len(all_events_expanded),  # Expanded event count
                    'analyzed_events': len(events_analysis),
                    'filtered_events': filtered_count,
                    'categories': dict(event_categories),
                    'category_filter_applied': bool(event_categories_to_process),
                    'categories_processed': event_categories_to_process,
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
        else:
            print(f"No events were successfully analyzed (filtered out {filtered_count} events)")
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'Found events but none were analyzed',
                    'events_found': items_count,
                    'expanded_events_found': len(all_events_expanded),
                    'filtered_events': filtered_count,
                    'category_filter_applied': bool(event_categories_to_process),
                    'categories_processed': event_categories_to_process,
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
            
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }



def is_org_view_enabled():
    """
    Check if AWS Health Organization View is enabled
    
    Returns:
        bool: True if organization view is enabled, False otherwise
    """
    try:
        # Try to call an organization-specific API to check if it's enabled
        health_client = boto3.client('health', region_name='us-east-1')
        # This will throw an exception if org view is not enabled
        health_client.describe_events_for_organization(
            filter={},
            maxResults=1
        )
        return True
    except Exception as e:
        error_code = getattr(e, 'response', {}).get('Error', {}).get('Code', '')
        if error_code == 'SubscriptionRequiredException':
            return False
        # For any other error, assume we don't have org view permissions
        return False

def get_bedrock_client():
    """
    Get Amazon Bedrock client
    
    Returns:
        boto3.client: Bedrock runtime client
    """
    # Use ap-southeast-1 for Bedrock as Claude 3.5 Sonnet v1 is available in this region
    return boto3.client(service_name='bedrock-runtime', region_name='ap-southeast-1')

def format_time(time_str):
    """
    Format time string to be consistent
    
    Args:
        time_str (str): ISO format time string
        
    Returns:
        str: Formatted time string (YYYY-MM-DD)
    """
    if not time_str or time_str == 'N/A':
        return 'N/A'
    
    try:
        # If it's already a datetime object
        if isinstance(time_str, datetime):
            return time_str.strftime('%Y-%m-%d')
        
        # Parse ISO format
        dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except Exception:
        # If we can't parse it, return as is
        return time_str

def fetch_health_event_details(event_arn):
    """
    Fetch detailed event information from AWS Health API
    
    Args:
        event_arn (str): ARN of the health event
        
    Returns:
        dict: Event details including affected resources
    """
    try:
        # Health API should go to us-east-1
        health_client = boto3.client('health', region_name='us-east-1')
        
        # Get event details
        event_details = health_client.describe_event_details(
            eventArns=[event_arn]
        )
        
        # Get affected entities
        affected_entities = health_client.describe_affected_entities(
            filter={
                'eventArns': [event_arn]
            }
        )
        
        return {
            'details': event_details.get('successfulSet', [{}])[0] if event_details.get('successfulSet') else {},
            'entities': affected_entities.get('entities', [])
        }
    except Exception as e:
        print(f"Error fetching Health API data: {str(e)}")
        return {'details': {}, 'entities': []}
         
def extract_affected_resources(entities):
    """
    Extract affected resources from Health API entities
    
    Args:
        entities (list): List of entity objects from Health API
        
    Returns:
        str: Comma-separated list of affected resources
    """
    if not entities:
        return "None specified"
    
    resources = []
    for entity in entities:
        entity_value = entity.get('entityValue', '')
        if entity_value:
            resources.append(entity_value)
    
    if resources:
        return ", ".join(resources)
    else:
        return "None specified"

def analyze_event_with_bedrock(bedrock_client, event_data):
    """
    Analyze an AWS Health event using Amazon Bedrock with focus on outage impact
    
    Args:
        bedrock_client: Amazon Bedrock client
        event_data (dict): Event data to analyze
        
    Returns:
        dict: Analyzed event data
    """
    try:
        # Get event details
        event_type = event_data.get('eventTypeCode', event_data.get('event_type', 'Unknown'))
        event_category = event_data.get('eventTypeCategory', event_data.get('event_type_category', 'Unknown'))
        region = event_data.get('region', 'Unknown')
        
        # Format start time if it's a datetime object
        start_time = event_data.get('startTime', event_data.get('start_time', 'Unknown'))
        if hasattr(start_time, 'isoformat'):
            start_time = start_time.isoformat()
        
        # Use description for analysis
        description = event_data.get('description', 'No description available')

        
        # Prepare prompt for Bedrock - ENHANCED FOR OUTAGE ANALYSIS
        print(f"Processing event: {event_type} with category {event_category}")
        print(f"Using description (length: {len(description)}): {description[:100]}...")
        
        prompt = f"""
        You are an AWS expert specializing in outage analysis and business continuity. Your task is to analyze this AWS Health event and determine its potential impact on workload availability, system connectivity, and service outages.
        
        AWS Health Event:
        - Type: {event_type}
        - Category: {event_category}
        - Region: {region}
        - Start Time: {start_time}
        
        Event Description:
        {description}
        
        IMPORTANT ANALYSIS FOCUS:
        1. Will this event cause workload downtime if required actions are not taken?
        2. Will there be any service outages associated with this event?
        3. Will the application/workload experience network integration issues between connecting systems?
        4. What specific AWS services or resources could be impacted?
         
        
        CRITICAL EVENT CRITERIA:
        - Any event that will cause service downtime should be marked as CRITICAL
        - Any event that will cause network integration or SSL issues between systems should be marked as CRITICAL
        - Any event that requires immediate action to prevent outage should be marked as URGENT time sensitivity
        - Events with high impact but no immediate downtime should be marked as HIGH risk level
    
        Please analyze this event and provide the following information in JSON format:
        {{
          "critical": boolean,
          "risk_level": "critical|high|medium|low",
          "account_impact": "critical|high|medium|low",
          "time_sensitivity": "Routine|Urgent|Critical",
          "risk_category": "Availability|Security|Performance|Cost|Compliance",
          "required_actions": "string",
          "impact_analysis": "string",
          "consequences_if_ignored": "string",
          "affected_resources": "string",
          "key_date": "YYYY-MM-DD or null"
        }}
        
        IMPORTANT: In your impact_analysis field, be very specific about:
        1. Potential outages and their estimated duration
        2. Connectivity issues between systems
        3. Whether this will cause downtime if actions are not taken
        
        In your consequences_if_ignored field, clearly state what outages or disruptions will occur if the event is not addressed.
        
        For the key_date field:
        - Analyze the event description for any dates that customers need to be aware of
        - Look for dates when actions must be taken, when changes will occur, or when impacts will begin
        - Return the EARLIEST date that will impact the customer in YYYY-MM-DD format
        - If no specific date is mentioned or can be determined, return null
        - Common date patterns to look for: deadlines, maintenance windows, deprecation dates, end-of-life dates, migration deadlines

        RISK LEVEL GUIDELINES:
        - CRITICAL: Will cause service outage or severe disruption if not addressed
        - HIGH: Significant impact but not an immediate outage
        - MEDIUM: Moderate impact requiring attention
        - LOW: Minimal impact, routine maintenance
        """
        
        # Determine which model we're using and format accordingly
        model_id = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-3-5-sonnet-20240620-v1:0')
        max_tokens = int(os.environ.get('BEDROCK_MAX_TOKENS', '4000'))
        temperature = float(os.environ.get('BEDROCK_TEMPERATURE', '0.2'))
        top_p = float(os.environ.get('BEDROCK_TOP_P', '0.9'))
        
        print(f"Sending request to Bedrock model: '{model_id}'")
        
        if "claude-3" in model_id.lower():
            # Claude 3 models use the messages format
            payload = {
                "modelId": model_id,
                "contentType": "application/json",
                "accept": "application/json",
                "body": json.dumps({
                    "anthropic_version": "bedrock-2023-05-31",
                    "max_tokens": max_tokens,
                    "temperature": temperature,
                    "top_p": top_p,
                    "messages": [
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                })
            }
        else:
            # Claude 2 and other models use the older prompt format
            payload = {
                "modelId": model_id,
                "contentType": "application/json",
                "accept": "application/json",
                "body": json.dumps({
                    "prompt": f"\n\nHuman: {prompt}\n\nAssistant:",
                    "max_tokens_to_sample": max_tokens,
                    "temperature": temperature,
                    "top_p": top_p
                })
            }
        
        # Call Bedrock
        try:
            response = bedrock_client.invoke_model(**payload)
            response_body = json.loads(response.get('body').read())
            
            # Extract response based on model
            if "claude-3" in model_id.lower():
                response_text = response_body.get('content', [{}])[0].get('text', '')
            else:
                response_text = response_body.get('completion', '')
            
            # Store the full analysis text as a string
            event_data['analysis_text'] = response_text
            
            # Try to extract JSON from the response
            json_match = re.search(r'```json\s*(.*?)\s*```', response_text, re.DOTALL)
            if json_match:
                json_str = json_match.group(1)
            else:
                json_match = re.search(r'({.*})', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(1)
                else:
                    json_str = response_text
            
            # Parse the JSON
            try:
                # Clean the JSON string to handle control characters from Claude 3.5 Sonnet v1
                cleaned_json = json_str.replace('\n', '\\n').replace('\t', '\\t').replace('\r', '\\r')
                # print(f"Attempting to parse JSON: {cleaned_json[:500]}...")
                analysis = json.loads(cleaned_json)
                # print(f"Successfully parsed JSON: {analysis}")
                
                # NEW CODE: Normalize risk level to ensure consistency
                if 'risk_level' in analysis:
                    risk_level = analysis['risk_level'].strip().upper()
                    
                    # Ensure "critical" is properly recognized and distinguished from "high"
                    if risk_level in ['CRITICAL', 'SEVERE']:
                        analysis['risk_level'] = 'CRITICAL'
                        # Make sure critical boolean flag is consistent
                        analysis['critical'] = True
                    elif risk_level == 'HIGH':
                        analysis['risk_level'] = 'HIGH'
                    elif risk_level in ['MEDIUM', 'MODERATE']:
                        analysis['risk_level'] = 'MEDIUM'
                    elif risk_level == 'LOW':
                        analysis['risk_level'] = 'LOW'
                    
                    # If critical flag is True but risk_level isn't CRITICAL, fix it
                    if analysis.get('critical', False) and analysis['risk_level'] != 'CRITICAL':
                        analysis['risk_level'] = 'CRITICAL'
                
                # Update event data with analysis
                event_data.update(analysis)
                
                return event_data
            except json.JSONDecodeError as e:
                print(f"JSON parsing failed: {str(e)}")
                # print(f"Failed to parse JSON from response: {json_str[:500]}...")
                # print(f"Full response length: {len(response_text)}")
                
                # Try to manually extract key fields if JSON parsing fails
                manual_analysis = {}
                
                # Extract critical status
                if '"critical": false' in response_text.lower():
                    manual_analysis['critical'] = False
                elif '"critical": true' in response_text.lower():
                    manual_analysis['critical'] = True
                
                # Extract risk level
                risk_match = re.search(r'"risk_level":\s*"([^"]+)"', response_text, re.IGNORECASE)
                if risk_match:
                    manual_analysis['risk_level'] = risk_match.group(1).lower()
                
                # Extract other fields
                for field in ['account_impact', 'time_sensitivity', 'risk_category', 'required_actions', 'impact_analysis', 'consequences_if_ignored', 'affected_resources']:
                    field_match = re.search(f'"({field})":\s*"([^"]+)"', response_text, re.IGNORECASE)
                    if field_match:
                        manual_analysis[field] = field_match.group(2)
                
                if manual_analysis:
                    # print(f"Manually extracted fields: {manual_analysis}")
                    event_data.update(manual_analysis)
                else:
                    # Provide default values if all parsing fails
                    event_data.update({
                        'critical': False,
                        'risk_level': 'low',
                        'account_impact': 'low',
                        'time_sensitivity': 'Routine',
                        'risk_category': 'Unknown',
                        'required_actions': 'Review event details manually',
                        'impact_analysis': 'Unable to automatically analyze this event',
                        'consequences_if_ignored': 'Unknown',
                        'affected_resources': 'Unknown'
                    })
                
                return event_data
                
        except Exception as e:
            print(f"Error in Bedrock analysis: {str(e)}")
            traceback.print_exc()
            
            # Provide default values if Bedrock analysis fails
            event_data.update({
                'critical': False,
                'risk_level': 'low',
                'account_impact': 'low',
                'time_sensitivity': 'Routine',
                'risk_category': 'Unknown',
                'required_actions': 'Review event details manually',
                'impact_analysis': 'Unable to automatically analyze this event',
                'consequences_if_ignored': 'Unknown',
                'affected_resources': 'Unknown',
                'analysis_text': f"Error during analysis: {str(e)}"
            })
            return event_data
    
    except Exception as e:
        print(f"Unexpected error in analyze_event_with_bedrock: {str(e)}")
        traceback.print_exc()
        
        # Provide default values if function fails
        event_data.update({
            'critical': False,
            'risk_level': 'low',
            'account_impact': 'low',
            'time_sensitivity': 'Routine',
            'risk_category': 'Unknown',
            'required_actions': 'Review event details manually',
            'impact_analysis': 'Unable to automatically analyze this event',
            'consequences_if_ignored': 'Unknown',
            'affected_resources': 'Unknown',
            'analysis_text': f"Error during analysis: {str(e)}"
        })
        return event_data



def categorize_analysis(analysis_text):
    """
    Extract structured data from Bedrock analysis text
    
    Args:
        analysis_text: Analysis text from Bedrock (string or dict)
        
    Returns:
        dict: Structured data extracted from analysis
    """
    categories = {
        'critical': False,
        'risk_level': 'low',
        'impact_analysis': '',
        'required_actions': '',
        'time_sensitivity': 'Routine',
        'risk_category': 'Unknown',
        'consequences_if_ignored': '',
        'event_category': 'Low',
        'key_date': 'N/A'
    }
    
    try:
        # If analysis_text is already a dictionary, use it directly
        if isinstance(analysis_text, dict):
            # Update our categories with values from the dictionary
            for key in categories.keys():
                if key in analysis_text:
                    categories[key] = analysis_text[key]
            
            # Also check for affected_resources
            if 'affected_resources' in analysis_text:
                categories['affected_resources'] = analysis_text['affected_resources']
                
            return categories
            
        # If analysis_text is not a string, convert it to string
        if not isinstance(analysis_text, str):
            analysis_text = str(analysis_text)
            
        # Try to parse as JSON first
        try:
            json_data = json.loads(analysis_text)
            # If successful, update our categories with values from the JSON
            for key in categories.keys():
                if key in json_data:
                    categories[key] = json_data[key]
            
            # Also check for affected_resources
            if 'affected_resources' in json_data:
                categories['affected_resources'] = json_data['affected_resources']
                
            return categories
        except json.JSONDecodeError:
            # Not valid JSON, continue with regex parsing
            pass
            
        # Extract critical status
        critical_match = re.search(r'CRITICAL:\s*(?:\[)?([Yy]es|[Nn]o)(?:\])?', analysis_text)
        if critical_match:
            categories['critical'] = critical_match.group(1).lower() == 'yes'
        
        # Extract risk level
        risk_match = re.search(r'RISK LEVEL:\s*(?:\[)?([Hh]igh|[Mm]edium|[Ll]ow)(?:\])?', analysis_text)
        if risk_match:
            categories['risk_level'] = risk_match.group(1).lower()
        
        # Extract account impact
        impact_match = re.search(r'ACCOUNT IMPACT:\s*(?:\[)?([Hh]igh|[Mm]edium|[Ll]ow)(?:\])?', analysis_text)
        if impact_match:
            categories['account_impact'] = impact_match.group(1).lower()
        
        # Extract impact analysis
        impact_analysis_match = re.search(r'IMPACT ANALYSIS:(.*?)(?:REQUIRED ACTIONS:|$)', analysis_text, re.DOTALL)
        if impact_analysis_match:
            categories['impact_analysis'] = impact_analysis_match.group(1).strip()
        
        # Extract required actions
        required_actions_match = re.search(r'REQUIRED ACTIONS:(.*?)(?:TIME SENSITIVITY:|$)', analysis_text, re.DOTALL)
        if required_actions_match:
            categories['required_actions'] = required_actions_match.group(1).strip()
        
        # Extract time sensitivity
        time_sensitivity_match = re.search(r'TIME SENSITIVITY:\s*([Ii]mmediate|[Uu]rgent|[Ss]oon|[Rr]outine)', analysis_text)
        if time_sensitivity_match:
            categories['time_sensitivity'] = time_sensitivity_match.group(1).capitalize()
        
        # Extract risk category
        risk_category_match = re.search(r'RISK CATEGORY:\s*([Tt]echnical|[Oo]perational|[Ss]ecurity|[Cc]ompliance|[Cc]ost|[Aa]vailability)', analysis_text)
        if risk_category_match:
            categories['risk_category'] = risk_category_match.group(1).capitalize()
        
        # Extract consequences if ignored
        consequences_match = re.search(r'CONSEQUENCES IF IGNORED:(.*?)(?:$)', analysis_text, re.DOTALL)
        if consequences_match:
            categories['consequences_if_ignored'] = consequences_match.group(1).strip()
        
        # Extract affected resources
        affected_match = re.search(r'AFFECTED RESOURCES:(.*?)(?:$)', analysis_text, re.DOTALL)
        if affected_match:
            categories['affected_resources'] = affected_match.group(1).strip()
            
    except Exception as e:
        print(f"Error categorizing analysis: {str(e)}")
    
    return categories


def create_excel_report_improved(events_analysis):
    """
    Create an improved Excel report with detailed event analysis
    
    Args:
        events_analysis (list): List of analyzed events data
        
    Returns:
        BytesIO: Excel file as bytes
    """
    # Create workbook and sheets
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    events_sheet = wb.create_sheet(title="All Events")
    
    # Add summary data
    summary_sheet['A1'] = "AWS Health Events Analysis Summary"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add event counts by category
    event_categories = defaultdict(int)
    for event in events_analysis:
        event_categories[event.get('event_type_category', 'unknown')] += 1
        if event.get('critical', False):
            event_categories['critical'] += 1
        
        risk_level = event.get('risk_level', 'low').lower()
        if risk_level == 'high':
            event_categories['high_risk'] += 1
        elif risk_level == 'medium':
            event_categories['medium_risk'] += 1
        elif risk_level == 'low':
            event_categories['low_risk'] += 1
            
        impact = event.get('account_impact', 'low').lower()
        if impact == 'high':
            event_categories['high_impact'] += 1
        elif impact == 'medium':
            event_categories['medium_impact'] += 1
        elif impact == 'low':
            event_categories['low_impact'] += 1
    
    # Add summary counts
    summary_sheet['A3'] = "Event Categories"
    summary_sheet['A3'].font = Font(bold=True)
    
    row = 4
    for category, count in event_categories.items():
        summary_sheet[f'A{row}'] = category.replace('_', ' ').title()
        summary_sheet[f'B{row}'] = count
        row += 1
    
    # Add headers to events sheet
    headers = [
        "Event ARN",  # Added Event ARN as first column
        "Event Type", 
        "Region", 
        "Start Time", 
        "Last Update", 
        "Category",
        "Description",  # Added Description column
        "Critical", 
        "Risk Level", 
        "Account ID", 
        "Key Date",  # Added Key Date column
        "Time Sensitivity", 
        "Risk Category", 
        "Required Actions", 
        "Impact Analysis", 
        "Consequences If Ignored", 
        "Affected Resources"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = events_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add event data
    for row_num, event in enumerate(events_analysis, 2):
        event_arn = event.get('eventArn', event.get('arn', 'N/A'))
        events_sheet.cell(row=row_num, column=1).value = event_arn # Add Event ARN
        events_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
        events_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
        events_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
        events_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
        events_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
        
        # Add description with text wrapping
        description_cell = events_sheet.cell(row=row_num, column=7)
        description_cell.value = event.get('description', 'N/A')
        description_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        events_sheet.cell(row=row_num, column=8).value = "Yes" if event.get('critical', False) else "No"
        events_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
        events_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
        events_sheet.cell(row=row_num, column=11).value = event.get('key_date', 'N/A')
        events_sheet.cell(row=row_num, column=12).value = event.get('time_sensitivity', 'Routine')
        events_sheet.cell(row=row_num, column=13).value = event.get('risk_category', 'Unknown')
        events_sheet.cell(row=row_num, column=14).value = event.get('required_actions', '')
        events_sheet.cell(row=row_num, column=15).value = event.get('impact_analysis', '')
        events_sheet.cell(row=row_num, column=16).value = event.get('consequences_if_ignored', '')
        events_sheet.cell(row=row_num, column=17).value = event.get('affected_resources', 'None')
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        if is_critical:
            for col_num in range(1, 18):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, 18):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, 18):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in events_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        events_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific width for description column
    events_sheet.column_dimensions['G'].width = 60  # Description column
    
    # Create critical events sheet
    critical_sheet = wb.create_sheet(title="Critical Events")
    
    # Add headers (same as events sheet)
    for col_num, header in enumerate(headers, 1):
        cell = critical_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add critical event data
    critical_events = [event for event in events_analysis if event.get('critical', False)]
    if not critical_events:
        critical_sheet.cell(row=2, column=1).value = "No critical events found"
        critical_sheet.merge_cells('A2:Q2')  # Updated to include new columns
        critical_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    else:
        for row_num, event in enumerate(critical_events, 2):
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            critical_sheet.cell(row=row_num, column=1).value = event_arn
            critical_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
            critical_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
            critical_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
            critical_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
            critical_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
            
            # Add description with text wrapping
            description_cell = critical_sheet.cell(row=row_num, column=7)
            description_cell.value = event.get('description', 'N/A')
            description_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            critical_sheet.cell(row=row_num, column=8).value = "Yes"
            critical_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
            critical_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
            critical_sheet.cell(row=row_num, column=11).value = event.get('key_date', 'N/A')
            critical_sheet.cell(row=row_num, column=12).value = event.get('time_sensitivity', 'Routine')
            critical_sheet.cell(row=row_num, column=13).value = event.get('risk_category', 'Unknown')
            critical_sheet.cell(row=row_num, column=14).value = event.get('required_actions', '')
            critical_sheet.cell(row=row_num, column=15).value = event.get('impact_analysis', '')
            critical_sheet.cell(row=row_num, column=16).value = event.get('consequences_if_ignored', '')
            critical_sheet.cell(row=row_num, column=17).value = event.get('affected_resources', 'None')
            
            # Apply critical highlighting
            for col_num in range(1, 18):
                critical_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
    
    # Auto-adjust column widths for critical sheet
    for col in critical_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        critical_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific width for description column in critical sheet
    critical_sheet.column_dimensions['G'].width = 60  # Description column
    
     # Create risk analysis sheet with full Bedrock analysis
    analysis_sheet = wb.create_sheet(title="Risk Analysis")
    
    # Add headers
    analysis_headers = ["Event Type", "Region", "Risk Level", "Full Analysis"]
    
    for col_num, header in enumerate(analysis_headers, 1):
        cell = analysis_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add analysis data
    for row_num, event in enumerate(events_analysis, 2):
        # Get event type from either format
        event_type = event.get('eventTypeCode', event.get('event_type', 'N/A'))
        analysis_sheet.cell(row=row_num, column=1).value = event_type
        
        # Get region
        analysis_sheet.cell(row=row_num, column=2).value = event.get('region', 'N/A')
        
        # Get risk level
        analysis_sheet.cell(row=row_num, column=3).value = event.get('risk_level', 'low').upper()
        
        # Fix: Ensure analysis_text is a string
        analysis_text = event.get('analysis_text', '')
        
        # Handle different data types
        if isinstance(analysis_text, dict):
            # Custom JSON encoder to handle datetime objects
            def datetime_handler(obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                return str(obj)
            
            try:
                analysis_text = json.dumps(analysis_text, indent=2, default=datetime_handler)
            except Exception as e:
                analysis_text = f"Error serializing analysis: {str(e)}"
        elif not isinstance(analysis_text, str):
            # Convert non-string values to string
            try:
                analysis_text = str(analysis_text)
            except Exception as e:
                analysis_text = f"Error converting to string: {str(e)}"
        
        # Set the cell value
        analysis_sheet.cell(row=row_num, column=4).value = analysis_text
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        if is_critical:
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    # Set column widths for analysis sheet
    analysis_sheet.column_dimensions['A'].width = 30
    analysis_sheet.column_dimensions['B'].width = 15
    analysis_sheet.column_dimensions['C'].width = 15
    analysis_sheet.column_dimensions['D'].width = 100
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)


    
    return excel_buffer

def generate_summary_html(total_events, event_categories, filtered_events, category_filter, events_analysis, accounts_processed=None):
    """
    Generate HTML summary for email
    
    Args:
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        filtered_events (int): Number of filtered events
        category_filter (list): Categories used for filtering
        events_analysis (list): Analyzed events data
        accounts_processed (str, optional): Comma-separated string of account IDs being processed
        
    Returns:
        str: HTML content for email
    """
    # Current date for the report
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    # Calculate accurate event counts directly from events_analysis
    critical_count = sum(1 for event in events_analysis if event.get('critical', False))
    high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
    medium_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'medium')
    low_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'low')

    # Define CSS for better table formatting
    table_css = """
    <style>
        .health-events-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 5px;
        }
        .health-events-table th, .health-events-table td {
            border: 1px solid #ddd;
            padding: 6px;
            text-align: left;
            vertical-align: top;
            word-wrap: break-word;
            font-size: 5px;
        }
        .health-events-table th {
            background-color: #f2f2f2;
        }
        /* Column width constraints */
        .col-arn {
            width: 15%;
        }
        .col-region {
            width: 8%;
        }
        .col-start-time {
            width: 12%;
        }
        .col-risk {
            width: 8%;
        }
        .col-accountid {
            width: 12%;
        }
    </style>
    """
    
    # Start building HTML content
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            .header {{ background-color: #232F3E; color: white; padding: 20px; }}
            .content {{ padding: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left;"font-size:5px;" }}
            th {{ background-color: #f2f2f2; }}
            .critical {{ background-color: #ffcccc; }}
            .high {{ background-color: #fff2cc; }}
            .medium {{ background-color: #e6f2ff; }}
            .summary {{ margin-top: 20px; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>AWS Health Events Analysis Report</h1>
            <p>Date: {current_date}</p>
        </div>
        <div class="content">
            <div class="summary">
                <h2>Summary</h2>
                <p>Total AWS Health events analyzed: {len(events_analysis)} of {total_events} events found</p>
    """
    
    # Add accounts processed information if provided
    if accounts_processed:
        html_content += f"""
                <p><strong>Accounts processed:</strong> {accounts_processed}</p>
        """
    
    # Add analysis window information
    start_time = datetime.now(timezone.utc) - timedelta(days=int(os.environ['ANALYSIS_WINDOW_DAYS']))
    html_content += f"""
                <p>Analysis Window: From {start_time.strftime('%Y-%m-%d %H:%M:%S')} UTC onwards (no end date)</p>
    """
    
    # Add filter information if applicable
    if category_filter:
        html_content += f"""
                <p>Events filtered by categories: {', '.join(category_filter)}</p>
                <p>Events excluded by filter: {filtered_events}</p>
        """
    
    # Add event category counts - USING ACCURATE COUNTS
    html_content += """
                <h3>Event Categories</h3>
                <ul>
    """
    
    # Add critical events count if any
    if critical_count > 0:
        html_content += f"""
                <li><strong>Critical Events:</strong> {critical_count}</li>
        """
    
    # Add risk level counts - USING ACCURATE COUNTS
    html_content += f"""
                <li><strong>High Risk Events:</strong> {high_risk_count}</li>
                <li><strong>Medium Risk Events:</strong> {medium_risk_count}</li>
                <li><strong>Low Risk Events:</strong> {low_risk_count}</li>
            </ul>
            </div>
    """
    
    # Add critical events table if any exist
    critical_events = [event for event in events_analysis if event.get('critical', False)]
    if critical_events:
        html_content += """
            <h2>Critical Events</h2>
            <table class="health-events-table">
                <tr>
                    <th style="font-size:12px;">Event ARN</th>
                    <th style="font-size:12px;">Region</th>
                    <th style="font-size:12px;">Start Time</th>
                    <th style="font-size:12px;">Risk Level</th>
                    <th style="font-size:12px;">Account ID</th>
                </tr>
        """
        
        for event in critical_events:
            # Get event ARN, preferring eventArn if available, falling back to arn
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            
            html_content += f"""
                <tr class="critical">
                    <td style="font-size:12px;">{event_arn}</td>
                    <td style="font-size:12px;">{event.get('region', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('start_time', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('risk_level', 'N/A').upper()}</td>
                    <td style="font-size:12px;">{event.get('accountId', 'N/A')}</td>
                </tr>
            """
        
        html_content += """
            </table>
        """
    
    # Add high risk events table
    high_risk_events = [event for event in events_analysis if event.get('risk_level', '').lower() == 'high']
    if high_risk_events:
        html_content += """
            <h2>High Risk Events</h2>
            <table class="health-events-table">
                <tr>
                    <th style="font-size:12px;">Event ARN</th>
                    <th style="font-size:12px;">Region</th>
                    <th style="font-size:12px;">Start Time</th>
                    <th style="font-size:12px;">Risk Level</th>
                    <th style="font-size:12px;">Account ID</th>
                </tr>
        """
        
        for event in high_risk_events:
            # Get event ARN, preferring eventArn if available, falling back to arn
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            
            html_content += f"""
                <tr class="high">
                    <td style="font-size:12px;">{event_arn}</td>
                    <td style="font-size:12px;">{event.get('region', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('start_time', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('risk_level', 'N/A').upper()}</td>
                    <td style="font-size:12px;">{event.get('accountId', 'N/A')}</td>
                </tr>
            """
        
        html_content += """
            </table>
        """
    
    # Add footer with attachment information
    html_content += """
            <div class="summary">
                <h2>Full Report</h2>
                <p>Please see the attached Excel file for complete details on all events.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content


def send_account_specific_emails(events_analysis, items_count, event_categories, filtered_count, event_categories_to_process):
    """
    Send emails grouped by recipient email address based on AWS Organizations account email addresses, 
    with SES failure handling and master email with tracking columns.
    
    Args:
        events_analysis (list): List of analyzed events
        items_count (int): Total number of original events
        event_categories (dict): Event categories count
        filtered_count (int): Number of filtered events
        event_categories_to_process (list): Categories that were processed
        
    Returns:
        None
    """
    try:
        # Get account email mapping using hierarchical approach (Parameter Store -> Organizations -> None)
        account_to_email, email_to_accounts = get_combined_account_email_mapping()
        
        # Initialize tracking for email sending results
        email_sending_results = {}  # email -> (success, error_message)
        events_for_default_email = []  # Events that need to go to default email
        
        # Add tracking columns to all events
        for event in events_analysis:
            account_id = event.get('accountId', 'N/A')
            if account_id in account_to_email:
                event['mapped_email'] = account_to_email[account_id]
                event['email_sent_status'] = 'Pending'
            else:
                event['mapped_email'] = 'No mapping'
                event['email_sent_status'] = 'Sent to default'
        
        if not account_to_email:
            # No account-specific mapping, send all events to default email
            print("No account email mapping found, sending all events to default email")
            # Update all events status
            for event in events_analysis:
                event['email_sent_status'] = 'Sent to default'
            
            excel_buffer = create_excel_report_improved_with_tracking(events_analysis, include_tracking_columns=True)
            summary_html = generate_summary_html(
                items_count, event_categories, filtered_count, 
                event_categories_to_process, events_analysis
            )
            send_ses_email_with_attachment(summary_html, excel_buffer, items_count, event_categories, events_analysis)
            return
        
        # Group events by email address (not account ID)
        events_by_email = defaultdict(list)
        unmapped_events = []
        
        for event in events_analysis:
            account_id = event.get('accountId', 'N/A')
            if account_id in account_to_email:
                email = account_to_email[account_id]
                events_by_email[email].append(event)
            else:
                unmapped_events.append(event)
                event['email_sent_status'] = 'Sent to default'
        
        print(f"Grouped events by email: {len(events_by_email)} unique email addresses, {len(unmapped_events)} unmapped events")
        
        # Send emails grouped by recipient email address with SES error handling
        for email, email_events in events_by_email.items():
            try:
                # Get the accounts that map to this email
                accounts_for_email = email_to_accounts[email]
                accounts_str = ', '.join(accounts_for_email)
                
                print(f"Attempting to send email to {email} for accounts [{accounts_str}] with {len(email_events)} events")
                
                # Create Excel report for all events going to this email with account filter info (no tracking columns)
                excel_buffer = create_excel_report_improved_with_tracking(email_events, accounts_str, include_tracking_columns=False)
                
                # Calculate event categories for this email's events
                email_event_categories = defaultdict(int)
                for event in email_events:
                    if event.get('critical', False):
                        email_event_categories['critical'] += 1
                    
                    risk_level = event.get('risk_level', 'low')
                    email_event_categories[f"{risk_level}_risk"] += 1
                    
                    account_impact = event.get('account_impact', 'low')
                    email_event_categories[f"{account_impact}_impact"] += 1
                
                # Generate summary HTML for this email's events with accounts information
                summary_html = generate_summary_html(
                    len(email_events), email_event_categories, 0, 
                    event_categories_to_process, email_events, accounts_str
                )
                
                # Check email verification status before attempting to send (only in sandbox mode)
                ses_client = boto3.client('ses')
                verification_failed = False
                verification_error = ""
                
                if should_verify_recipient_email():
                    print(f"SES is in sandbox mode - verifying recipient email: {email}")
                    try:
                        identity_response = ses_client.get_identity_verification_attributes(
                            Identities=[email]
                        )
                        
                        verification_status = identity_response.get('VerificationAttributes', {}).get(
                            email, {}
                        ).get('VerificationStatus', 'NotStarted')
                        
                        if verification_status != 'Success':
                            verification_failed = True
                            verification_error = f"Recipient email not verified (status: {verification_status})"
                            print(f"Email verification failed for {email}: {verification_error}")
                            
                    except Exception as e:
                        print(f"Warning: Could not verify recipient status for {email}: {e}")
                        # Continue with sending - this is just a warning
                else:
                    print(f"SES is in production mode - skipping verification check for: {email}")
                
                # Update event status based on verification check
                if verification_failed:
                    # Mark events as failed due to verification
                    for event in email_events:
                        event['email_sent_status'] = f'Failed: {verification_error}'
                    
                    # Add these events to default email
                    events_for_default_email.extend(email_events)
                    
                    # Track the verification failure
                    email_sending_results[email] = (False, verification_error)
                    
                else:
                    # Attempt to send email with SES error handling
                    success, error_message = send_ses_email_with_attachment_to_recipient_with_error_handling(
                        summary_html, excel_buffer, len(email_events), 
                        email_event_categories, email_events, 
                        email, accounts_str
                    )
                    
                    # Track the result
                    email_sending_results[email] = (success, error_message)
                    
                    if success:
                        print(f"Successfully sent email to {email}")
                        # Update event status for successful sends
                        for event in email_events:
                            event['email_sent_status'] = 'Sent successfully'
                    else:
                        print(f"Failed to send email to {email}: {error_message}")
                        # Add these events to default email and update status
                        events_for_default_email.extend(email_events)
                        for event in email_events:
                            event['email_sent_status'] = f'Failed: {error_message}'
                
            except Exception as e:
                error_msg = str(e)
                print(f"Unexpected error sending email to {email}: {error_msg}")
                traceback.print_exc()
                
                # Track the failure
                email_sending_results[email] = (False, error_msg)
                
                # Add these events to default email and update status
                events_for_default_email.extend(email_events)
                for event in email_events:
                    event['email_sent_status'] = f'Failed: {error_msg}'
        
        # Log summary of what happened
        total_failed_events = len(events_for_default_email)
        total_unmapped_events = len(unmapped_events)
        print(f"Email sending summary: {len(events_for_default_email)} failed events, {len(unmapped_events)} unmapped events")
        
        # ALWAYS send master email with all events and tracking columns (this is the ONLY email to default recipients)
        try:
            print(f"Sending master email with all {len(events_analysis)} events and tracking information")
            
            # Create master Excel report with tracking columns
            master_excel_buffer = create_excel_report_improved_with_tracking(events_analysis, include_tracking_columns=True)
            
            # Generate master summary HTML
            master_summary_html = generate_summary_html(
                items_count, event_categories, filtered_count, 
                event_categories_to_process, events_analysis
            )
            
            # Send master email to default recipients
            send_ses_email_with_attachment_master(master_summary_html, master_excel_buffer, items_count, event_categories, events_analysis)
            
        except Exception as e:
            print(f"Error sending master email: {str(e)}")
            traceback.print_exc()
            
    except Exception as e:
        print(f"Error in send_account_specific_emails: {str(e)}")
        traceback.print_exc()
        # Fallback to sending all events to default email
        try:
            # Add tracking columns for fallback
            for event in events_analysis:
                if 'mapped_email' not in event:
                    event['mapped_email'] = 'No mapping'
                if 'email_sent_status' not in event:
                    event['email_sent_status'] = 'Sent to default (fallback)'
            
            excel_buffer = create_excel_report_improved_with_tracking(events_analysis, include_tracking_columns=True)
            summary_html = generate_summary_html(
                items_count, event_categories, filtered_count, 
                event_categories_to_process, events_analysis
            )
            send_ses_email_with_attachment(summary_html, excel_buffer, items_count, event_categories, events_analysis)
        except Exception as fallback_error:
            print(f"Error in fallback email sending: {str(fallback_error)}")


def send_ses_email_with_attachment_to_recipient_with_error_handling(html_content, excel_buffer, total_events, event_categories, events_analysis, recipient_email, accounts_str):
    """
    Send email with Excel attachment to a specific recipient with SES error handling
    
    Args:
        html_content (str): HTML content for email body
        excel_buffer (BytesIO): Excel file as bytes
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        events_analysis (list): List of analyzed events
        recipient_email (str): Email address to send to
        accounts_str (str): Comma-separated string of account IDs for subject line and filename
        
    Returns:
        tuple: (success: bool, error_message: str)
    """
    try:
        # Get email configuration from environment variables
        sender = os.environ['SENDER_EMAIL']
        recipients = [recipient_email.strip()]
        
        # Create email subject
        critical_count = event_categories.get('critical', 0)
        high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
        
        if critical_count > 0:
            subject = f"{customer_name} [CRITICAL] AWS Health Events Analysis - {critical_count} Critical, {high_risk_count} High Risk Events"
        elif high_risk_count > 0:
            subject = f"{customer_name} [HIGH RISK] AWS Health Events Analysis - {high_risk_count} High Risk Events"
        else:
            subject = f"{customer_name} AWS Health Events Analysis - {total_events} Events"
        
        # Generate Excel filename
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H-%M')
        excel_filename = f"AWS_Health_Events_Analysis_{current_date}_{current_time}.xlsx"
        
        # Create SES client
        ses_client = boto3.client('ses')
        
        # Create raw email message with attachment
        msg_raw = {
            'Source': sender,
            'Destinations': recipients,
            'RawMessage': {
                'Data': create_raw_email_with_attachment(
                    sender=sender,
                    recipients=recipients,
                    subject=subject,
                    html_body=html_content,
                    attachment_data=excel_buffer.getvalue(),
                    attachment_name=excel_filename
                )
            }
        }
        
        # Check if recipient email is verified before sending (only in sandbox mode)
        if should_verify_recipient_email():
            print(f"SES is in sandbox mode - verifying recipient email: {recipient_email}")
            try:
                identity_response = ses_client.get_identity_verification_attributes(
                    Identities=[recipient_email]
                )
                
                verification_status = identity_response.get('VerificationAttributes', {}).get(
                    recipient_email, {}
                ).get('VerificationStatus', 'NotStarted')
                
                if verification_status != 'Success':
                    return False, f"Recipient email not verified (status: {verification_status})"
                    
            except Exception as e:
                print(f"Warning: Could not verify recipient status for {recipient_email}: {e}")
                # Continue with sending - this is just a warning, don't fail the send
        else:
            print(f"SES is in production mode - skipping verification check for: {recipient_email}")
        
        # Send email with SES error handling
        response = ses_client.send_raw_email(**msg_raw)
        print(f"Account-specific email sent successfully to {recipient_email}. Message ID: {response['MessageId']}")

        # Upload to S3 if configured
        try:
            # Determine which bucket to use - external or internal
            bucket_name = S3_BUCKET_NAME if S3_BUCKET_NAME else REPORTS_BUCKET
            
            if bucket_name:
                # Create S3 client
                s3_client = boto3.client('s3')
                
                # Generate S3 key with prefix if provided (only for external bucket)
                if S3_BUCKET_NAME:
                    s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{excel_filename}" if S3_KEY_PREFIX else excel_filename
                else:
                    s3_key = excel_filename  # No prefix for internal bucket
                
                # Reset buffer position to the beginning
                excel_buffer.seek(0)
                
                # Upload buffer to S3 using put_object
                s3_client.put_object(
                    Bucket=bucket_name,
                    Key=s3_key,
                    Body=excel_buffer.getvalue(),
                    ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Generate S3 URL for the uploaded file
                s3_url = f"s3://{bucket_name}/{s3_key}"
                print(f"Successfully uploaded account-specific file to {s3_url}")
                     
        except Exception as e:
            print(f"Error uploading account-specific file to S3: {str(e)}")
        
        return True, "Success"
        
    except ClientError as e:
        # Handle SES-specific errors
        error_code = e.response['Error']['Code']
        error_message = e.response['Error']['Message']
        
        if error_code == 'MessageRejected':
            if 'not verified' in error_message.lower() or 'pending approval' in error_message.lower():
                return False, f"Email not verified: {error_message}"
            else:
                return False, f"Message rejected: {error_message}"
        elif error_code == 'SendingPausedException':
            return False, f"Sending paused: {error_message}"
        elif error_code == 'MailFromDomainNotVerifiedException':
            return False, f"Domain not verified: {error_message}"
        else:
            return False, f"SES error ({error_code}): {error_message}"
            
    except Exception as e:
        return False, f"Unexpected error: {str(e)}"


def send_ses_email_with_attachment_master(html_content, excel_buffer, total_events, event_categories, events_analysis):
    """
    Send master email with Excel attachment containing all events and tracking columns
    
    Args:
        html_content (str): HTML content for email body
        excel_buffer (BytesIO): Excel file as bytes
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        events_analysis (list): List of analyzed events
        
    Returns:
        None
    """
    try:
        # Get email configuration from environment variables
        sender = os.environ['SENDER_EMAIL']
        recipients_str = os.environ['RECIPIENT_EMAILS']
        recipients = [email.strip() for email in recipients_str.split(',')]
        
        # Create email subject with counts
        critical_count = event_categories.get('critical', 0)
        high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
        
        if critical_count > 0:
            subject = f"{customer_name} [MASTER] [CRITICAL] AWS Health Events Analysis - {critical_count} Critical, {high_risk_count} High Risk Events"
        elif high_risk_count > 0:
            subject = f"{customer_name} [MASTER] [HIGH RISK] AWS Health Events Analysis - {high_risk_count} High Risk Events"
        else:
            subject = f"{customer_name} [MASTER] AWS Health Events Analysis - {total_events} Events"
        
        # Generate Excel filename
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H-%M')
        excel_filename = f"AWS_Health_Events_Analysis_MASTER_{current_date}_{current_time}.xlsx"
        
        # Create SES client
        ses_client = boto3.client('ses')
        
        # Create raw email message with attachment
        msg_raw = {
            'Source': sender,
            'Destinations': recipients,
            'RawMessage': {
                'Data': create_raw_email_with_attachment(
                    sender=sender,
                    recipients=recipients,
                    subject=subject,
                    html_body=html_content,
                    attachment_data=excel_buffer.getvalue(),
                    attachment_name=excel_filename
                )
            }
        }
        
        # Check if recipient emails are verified before sending (only in sandbox mode)
        unverified_recipients = []
        if should_verify_recipient_email():
            print("SES is in sandbox mode - verifying master email recipients")
            for recipient in recipients:
                try:
                    identity_response = ses_client.get_identity_verification_attributes(
                        Identities=[recipient]
                    )
                    
                    verification_status = identity_response.get('VerificationAttributes', {}).get(
                        recipient, {}
                    ).get('VerificationStatus', 'NotStarted')
                    
                    if verification_status != 'Success':
                        unverified_recipients.append(f"{recipient} (status: {verification_status})")
                        
                except Exception as e:
                    print(f"Warning: Could not verify recipient status for {recipient}: {e}")
                    # Continue with sending - this is just a warning
            
            if unverified_recipients:
                print(f"Warning: Some master email recipients are not verified: {', '.join(unverified_recipients)}")
        else:
            print("SES is in production mode - skipping verification checks for master email recipients")
        
        # Send email
        response = ses_client.send_raw_email(**msg_raw)
        print(f"Master email sent successfully. Message ID: {response['MessageId']}")

        # Upload to S3 if configured
        try:
            # Determine which bucket to use - external or internal
            bucket_name = S3_BUCKET_NAME if S3_BUCKET_NAME else REPORTS_BUCKET
            
            if bucket_name:
                # Create S3 client
                s3_client = boto3.client('s3')
                
                # Generate S3 key with prefix if provided (only for external bucket)
                if S3_BUCKET_NAME:
                    s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{excel_filename}" if S3_KEY_PREFIX else excel_filename
                else:
                    s3_key = excel_filename  # No prefix for internal bucket
                
                # Reset buffer position to the beginning
                excel_buffer.seek(0)
                
                # Upload buffer to S3 using put_object
                s3_client.put_object(
                    Bucket=bucket_name,
                    Key=s3_key,
                    Body=excel_buffer.getvalue(),
                    ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Generate S3 URL for the uploaded file
                s3_url = f"s3://{bucket_name}/{s3_key}"
                print(f"Successfully uploaded master file to {s3_url}")
                     
        except Exception as e:
            print(f"Error uploading master file to S3: {str(e)}")
        
    except Exception as e:
        print(f"Error sending master email: {str(e)}")
        traceback.print_exc()


def get_account_mapping_with_sources(include_availability_status=False):
    """
    Get account-email mapping with detailed source tracking for each account.
    
    Args:
        include_availability_status (bool): Whether to include availability status when both sources are enabled
    
    Returns:
        list: List of dictionaries with keys: account_id, email, source, availability_status (optional)
    """
    mapping_data = []
    
    try:
        # Get Organizations mapping first
        org_account_to_email = {}
        if USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING:
            org_account_to_email, _ = get_organization_account_email_mapping()
        
        # Get DynamoDB mapping
        custom_account_to_email = {}
        if USE_CUSTOM_ACCOUNT_EMAIL_MAPPING:
            custom_account_to_email, _ = get_custom_account_email_mapping_from_dynamodb()
        
        # Combine all accounts from both sources
        all_accounts = set()
        if org_account_to_email:
            all_accounts.update(org_account_to_email.keys())
        if custom_account_to_email:
            all_accounts.update(custom_account_to_email.keys())
        
        # Determine source and email for each account
        for account_id in all_accounts:
            email = None
            source = None
            availability_status = None
            
            # Check if account exists in DynamoDB (highest priority)
            if account_id in custom_account_to_email:
                email = custom_account_to_email[account_id]
                source = "Custom DynamoDB"
            # Fallback to Organizations mapping
            elif account_id in org_account_to_email:
                email = org_account_to_email[account_id]
                source = "AWS Organizations"
            
            # Determine availability status when both sources are enabled and requested
            if include_availability_status and USE_CUSTOM_ACCOUNT_EMAIL_MAPPING and USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING:
                in_dynamodb = account_id in custom_account_to_email
                in_organizations = account_id in org_account_to_email
                
                if in_dynamodb and not in_organizations:
                    availability_status = "DynamoDB only"
                elif in_organizations and not in_dynamodb:
                    availability_status = "Organizations only"
                elif in_dynamodb and in_organizations:
                    availability_status = "Both sources"
                else:
                    availability_status = "Unknown"  # Should not happen
            
            if email and source:
                item = {
                    'account_id': account_id,
                    'email': email,
                    'source': source
                }
                
                if include_availability_status:
                    item['availability_status'] = availability_status
                
                mapping_data.append(item)
        
        print(f"Generated mapping data with sources for {len(mapping_data)} accounts")
        
        # Log source breakdown
        source_counts = {}
        for item in mapping_data:
            source_counts[item['source']] = source_counts.get(item['source'], 0) + 1
        
        for source, count in source_counts.items():
            print(f"  {source}: {count} accounts")
        
        # Log availability status breakdown if included
        if include_availability_status and USE_CUSTOM_ACCOUNT_EMAIL_MAPPING and USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING:
            status_counts = {}
            for item in mapping_data:
                status = item.get('availability_status', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            
            print("  Availability status breakdown:")
            for status, count in status_counts.items():
                print(f"    {status}: {count} accounts")
        
    except Exception as e:
        print(f"Error generating mapping data with sources: {str(e)}")
        traceback.print_exc()
    
    return mapping_data


def create_excel_report_improved_with_tracking(events_analysis, account_filter_info=None, include_tracking_columns=True):
    """
    Create an improved Excel report with detailed event analysis and optional tracking columns
    
    Args:
        events_analysis (list): List of analyzed events data
        account_filter_info (str, optional): Comma-separated account IDs if filtered for specific accounts
        include_tracking_columns (bool): Whether to include "Mapped Email" and "Email Status" columns
        
    Returns:
        BytesIO: Excel file as bytes
    """
    # Create workbook and sheets
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    events_sheet = wb.create_sheet(title="All Events")
    
    # Add summary data
    summary_sheet['A1'] = "AWS Health Events Analysis Summary"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add event counts by category
    event_categories = defaultdict(int)
    for event in events_analysis:
        event_categories[event.get('event_type_category', 'unknown')] += 1
        if event.get('critical', False):
            event_categories['critical'] += 1
        
        risk_level = event.get('risk_level', 'low').lower()
        if risk_level == 'high':
            event_categories['high_risk'] += 1
        elif risk_level == 'medium':
            event_categories['medium_risk'] += 1
        elif risk_level == 'low':
            event_categories['low_risk'] += 1
            
        impact = event.get('account_impact', 'low').lower()
        if impact == 'high':
            event_categories['high_impact'] += 1
        elif impact == 'medium':
            event_categories['medium_impact'] += 1
        elif impact == 'low':
            event_categories['low_impact'] += 1
    
    # Add summary counts
    summary_sheet['A3'] = "Event Categories"
    summary_sheet['A3'].font = Font(bold=True)
    
    row = 4
    for category, count in event_categories.items():
        summary_sheet[f'A{row}'] = category.replace('_', ' ').title()
        summary_sheet[f'B{row}'] = count
        row += 1
    
    # Removed: Filtered Account IDs section - now using Account Email Mapping sheet instead
    
    # Add configuration summary for master reports only (when tracking columns are included)
    if include_tracking_columns:
        try:
            summary_sheet['A9'] = "Configuration Summary:"
            summary_sheet['A9'].font = Font(bold=True)
            
            config_row = 10
            summary_sheet[f'A{config_row}'] = f"Custom DynamoDB Mapping: {'ENABLED' if USE_CUSTOM_ACCOUNT_EMAIL_MAPPING else 'DISABLED'}"
            config_row += 1
            summary_sheet[f'A{config_row}'] = f"AWS Organizations Mapping: {'ENABLED' if USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING else 'DISABLED'}"
            config_row += 1
            
            if USE_CUSTOM_ACCOUNT_EMAIL_MAPPING:
                table_name = os.environ.get('ACCOUNT_EMAIL_MAPPING_TABLE', 'Not configured')
                summary_sheet[f'A{config_row}'] = f"DynamoDB Table: {table_name}"
                config_row += 1
            
            # Add default recipients info
            default_recipients = os.environ.get('RECIPIENT_EMAILS', 'Not configured')
            summary_sheet[f'A{config_row}'] = f"Default Recipients: {default_recipients}"
            
        except Exception as e:
            print(f"Error adding configuration summary to Summary sheet: {str(e)}")
            summary_sheet['A9'] = f"Configuration Summary: Error - {str(e)}"
    
    # Add headers to events sheet with optional tracking columns
    headers = [
        "Event ARN",
        "Event Type", 
        "Region", 
        "Start Time", 
        "Last Update", 
        "Category",
        "Description",
        "Critical", 
        "Risk Level", 
        "Account ID", 
        "Key Date",
        "Time Sensitivity", 
        "Risk Category", 
        "Required Actions", 
        "Impact Analysis", 
        "Consequences If Ignored", 
        "Affected Resources"
    ]
    
    # Add tracking columns only if requested (for master email)
    if include_tracking_columns:
        headers.extend([
            "Mapped Email",  # Tracking column
            "Email Status"   # Tracking column
        ])
    
    for col_num, header in enumerate(headers, 1):
        cell = events_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add event data with tracking columns
    for row_num, event in enumerate(events_analysis, 2):
        event_arn = event.get('eventArn', event.get('arn', 'N/A'))
        events_sheet.cell(row=row_num, column=1).value = event_arn
        events_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
        events_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
        events_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
        events_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
        events_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
        
        # Add description with text wrapping
        description_cell = events_sheet.cell(row=row_num, column=7)
        description_cell.value = event.get('description', 'N/A')
        description_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        events_sheet.cell(row=row_num, column=8).value = "Yes" if event.get('critical', False) else "No"
        events_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
        events_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
        events_sheet.cell(row=row_num, column=11).value = event.get('key_date', 'N/A')
        events_sheet.cell(row=row_num, column=12).value = event.get('time_sensitivity', 'Routine')
        events_sheet.cell(row=row_num, column=13).value = event.get('risk_category', 'Unknown')
        events_sheet.cell(row=row_num, column=14).value = event.get('required_actions', '')
        events_sheet.cell(row=row_num, column=15).value = event.get('impact_analysis', '')
        events_sheet.cell(row=row_num, column=16).value = event.get('consequences_if_ignored', '')
        events_sheet.cell(row=row_num, column=17).value = event.get('affected_resources', 'None')
        
        # Add tracking columns only if requested
        if include_tracking_columns:
            events_sheet.cell(row=row_num, column=18).value = event.get('mapped_email', 'No mapping')
            events_sheet.cell(row=row_num, column=19).value = event.get('email_sent_status', 'Unknown')
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        # Determine the number of columns based on whether tracking columns are included
        max_col = 19 if include_tracking_columns else 17
        
        if is_critical:
            for col_num in range(1, max_col + 1):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, max_col + 1):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, max_col + 1):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in events_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        events_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific widths for key columns
    events_sheet.column_dimensions['G'].width = 60  # Description column
    
    # Set tracking column widths only if they exist
    if include_tracking_columns:
        events_sheet.column_dimensions['R'].width = 30  # Mapped Email column
        events_sheet.column_dimensions['S'].width = 40  # Email Status column
    
    # Create critical events sheet with tracking columns
    critical_sheet = wb.create_sheet(title="Critical Events")
    
    # Add headers (same as events sheet)
    for col_num, header in enumerate(headers, 1):
        cell = critical_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add critical event data
    critical_events = [event for event in events_analysis if event.get('critical', False)]
    if not critical_events:
        critical_sheet.cell(row=2, column=1).value = "No critical events found"
        # Merge cells based on whether tracking columns are included
        merge_range = 'A2:S2' if include_tracking_columns else 'A2:Q2'
        critical_sheet.merge_cells(merge_range)
        critical_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    else:
        for row_num, event in enumerate(critical_events, 2):
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            critical_sheet.cell(row=row_num, column=1).value = event_arn
            critical_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
            critical_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
            critical_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
            critical_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
            critical_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
            
            # Add description with text wrapping
            description_cell = critical_sheet.cell(row=row_num, column=7)
            description_cell.value = event.get('description', 'N/A')
            description_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            critical_sheet.cell(row=row_num, column=8).value = "Yes"
            critical_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
            critical_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
            critical_sheet.cell(row=row_num, column=11).value = event.get('key_date', 'N/A')
            critical_sheet.cell(row=row_num, column=12).value = event.get('time_sensitivity', 'Routine')
            critical_sheet.cell(row=row_num, column=13).value = event.get('risk_category', 'Unknown')
            critical_sheet.cell(row=row_num, column=14).value = event.get('required_actions', '')
            critical_sheet.cell(row=row_num, column=15).value = event.get('impact_analysis', '')
            critical_sheet.cell(row=row_num, column=16).value = event.get('consequences_if_ignored', '')
            critical_sheet.cell(row=row_num, column=17).value = event.get('affected_resources', 'None')
            
            # Add tracking columns only if requested
            if include_tracking_columns:
                critical_sheet.cell(row=row_num, column=18).value = event.get('mapped_email', 'No mapping')
                critical_sheet.cell(row=row_num, column=19).value = event.get('email_sent_status', 'Unknown')
            
            # Apply critical highlighting
            max_col = 19 if include_tracking_columns else 17
            for col_num in range(1, max_col + 1):
                critical_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
    
    # Auto-adjust column widths for critical sheet
    for col in critical_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        critical_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific widths for key columns in critical sheet
    critical_sheet.column_dimensions['G'].width = 60  # Description column
    
    # Set tracking column widths only if they exist
    if include_tracking_columns:
        critical_sheet.column_dimensions['R'].width = 30  # Mapped Email column
        critical_sheet.column_dimensions['S'].width = 40  # Email Status column
    
    # Create risk analysis sheet with full Bedrock analysis
    analysis_sheet = wb.create_sheet(title="Risk Analysis")
    
    # Add headers
    analysis_headers = ["Event Type", "Region", "Risk Level", "Full Analysis"]
    
    for col_num, header in enumerate(analysis_headers, 1):
        cell = analysis_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add analysis data
    for row_num, event in enumerate(events_analysis, 2):
        # Get event type from either format
        event_type = event.get('eventTypeCode', event.get('event_type', 'N/A'))
        analysis_sheet.cell(row=row_num, column=1).value = event_type
        
        # Get region
        analysis_sheet.cell(row=row_num, column=2).value = event.get('region', 'N/A')
        
        # Get risk level
        analysis_sheet.cell(row=row_num, column=3).value = event.get('risk_level', 'low').upper()
        
        # Fix: Ensure analysis_text is a string
        analysis_text = event.get('analysis_text', '')
        
        # Handle different data types
        if isinstance(analysis_text, dict):
            # Custom JSON encoder to handle datetime objects
            def datetime_handler(obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                return str(obj)
            
            try:
                analysis_text = json.dumps(analysis_text, indent=2, default=datetime_handler)
            except Exception as e:
                analysis_text = f"Error serializing analysis: {str(e)}"
        elif not isinstance(analysis_text, str):
            # Convert non-string values to string
            try:
                analysis_text = str(analysis_text)
            except Exception as e:
                analysis_text = f"Error converting to string: {str(e)}"
        
        # Set the cell value
        analysis_sheet.cell(row=row_num, column=4).value = analysis_text
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        if is_critical:
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    
    # Set column widths for analysis sheet
    analysis_sheet.column_dimensions['A'].width = 30
    analysis_sheet.column_dimensions['B'].width = 15
    analysis_sheet.column_dimensions['C'].width = 15
    analysis_sheet.column_dimensions['D'].width = 100
    
    # Add account-email mapping configuration sheet for transparency (all reports)
    # This provides visibility into account-email routing for both master and account-specific reports
    mapping_sheet = wb.create_sheet(title="Account Email Mapping")
    
    # Determine if we should include availability status column (master reports only with both sources enabled)
    include_availability_column = (include_tracking_columns and 
                                 USE_CUSTOM_ACCOUNT_EMAIL_MAPPING and 
                                 USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING)
    
    # Add headers for mapping sheet
    mapping_sheet['A1'] = "Account ID"
    mapping_sheet['B1'] = "Email Address"
    mapping_sheet['C1'] = "Mapping Source"
    
    # Add availability status column header if needed
    header_cols = ['A1', 'B1', 'C1']
    if include_availability_column:
        mapping_sheet['D1'] = "Availability Status"
        header_cols.append('D1')
    
    # Style headers
    for col in header_cols:
        mapping_sheet[col].font = Font(bold=True)
        mapping_sheet[col].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        mapping_sheet[col].font = Font(bold=True, color="FFFFFF")
    
    # Get current account mapping configuration with source tracking
    try:
        account_to_email, email_to_accounts = get_combined_account_email_mapping()
        
        # Get detailed mapping with source information (include availability status for master reports with both sources)
        mapping_data = get_account_mapping_with_sources(include_availability_status=include_availability_column)
        
        # Add mapping data to sheet
        row = 2
        if mapping_data:
            # Filter mappings for account-specific reports
            if account_filter_info and not include_tracking_columns:
                # For account-specific reports, only show mappings for accounts in this report
                filtered_account_ids = set(acc.strip() for acc in account_filter_info.split(','))
                mapping_data = [item for item in mapping_data if item['account_id'] in filtered_account_ids]
            
            # Sort by email address for master reports, by account ID for account-specific reports
            if include_tracking_columns:
                # Master report: sort by email address, then by account ID
                mapping_data.sort(key=lambda x: (x['email'], x['account_id']))
            else:
                # Account-specific report: sort by account ID
                mapping_data.sort(key=lambda x: x['account_id'])
            
            # Add data to sheet
            for item in mapping_data:
                mapping_sheet[f'A{row}'] = item['account_id']
                mapping_sheet[f'B{row}'] = item['email']
                mapping_sheet[f'C{row}'] = item['source']
                
                # Add availability status column if included
                if include_availability_column:
                    mapping_sheet[f'D{row}'] = item.get('availability_status', 'Unknown')
                
                row += 1
        else:
            # No mappings configured
            mapping_sheet['A2'] = "No account-email mappings configured"
            mapping_sheet['B2'] = "All events sent to default recipients"
            mapping_sheet['C2'] = "Default configuration"
            
            # Merge cells appropriately based on number of columns
            if include_availability_column:
                mapping_sheet.merge_cells('A2:D2')
            else:
                mapping_sheet.merge_cells('A2:C2')
            row = 3
        
    except Exception as e:
        print(f"Error adding mapping sheet: {str(e)}")
        # Add error information to sheet
        mapping_sheet['A2'] = "Error retrieving mapping configuration"
        mapping_sheet['B2'] = str(e)
        mapping_sheet['C2'] = "Error"
    
    # Set column widths for mapping sheet
    mapping_sheet.column_dimensions['A'].width = 20
    mapping_sheet.column_dimensions['B'].width = 40
    mapping_sheet.column_dimensions['C'].width = 25
    
    # Set width for availability status column if included
    if include_availability_column:
        mapping_sheet.column_dimensions['D'].width = 20
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer



def send_ses_email_with_attachment(html_content, excel_buffer, total_events, event_categories,events_analysis):
    """
    Send email with Excel attachment using Amazon SES
    
    Args:
        html_content (str): HTML content for email body
        excel_buffer (BytesIO): Excel file as bytes
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        
    Returns:
        None
    """
    try:
        # Get email configuration from environment variables
        sender = os.environ['SENDER_EMAIL']
        recipients_str = os.environ['RECIPIENT_EMAILS']
        recipients = [email.strip() for email in recipients_str.split(',')]
        
        # Create email subject with counts
        critical_count = event_categories.get('critical', 0)
        high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
        
        if critical_count > 0:
            subject = f"{customer_name} [CRITICAL] AWS Health Events Analysis - {critical_count} Critical, {high_risk_count} High Risk Events"
        elif high_risk_count > 0:
            subject = f"{customer_name} [HIGH RISK] AWS Health Events Analysis - {high_risk_count} High Risk Events"
        else:
            subject = f"{customer_name} AWS Health Events Analysis - {total_events} Events"
        
        # Generate Excel filename
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H-%M')
        excel_filename = EXCEL_FILENAME_TEMPLATE.format(date=current_date, time=current_time)
        
        # Create SES client
        ses_client = boto3.client('ses')
        
        # Check if recipient emails are verified before sending (only in sandbox mode)
        if should_verify_recipient_email():
            print("SES is in sandbox mode - verifying recipient emails")
            unverified_recipients = []
            for recipient in recipients:
                try:
                    identity_response = ses_client.get_identity_verification_attributes(
                        Identities=[recipient]
                    )
                    
                    verification_status = identity_response.get('VerificationAttributes', {}).get(
                        recipient, {}
                    ).get('VerificationStatus', 'NotStarted')
                    
                    if verification_status != 'Success':
                        unverified_recipients.append(f"{recipient} (status: {verification_status})")
                        
                except Exception as e:
                    print(f"Warning: Could not verify recipient status for {recipient}: {e}")
            
            if unverified_recipients:
                print(f"Warning: Some recipients are not verified: {', '.join(unverified_recipients)}")
                print("Email sending may fail for unverified recipients in sandbox mode")
        else:
            print("SES is in production mode - skipping verification checks for recipients")
        
        # Create message container
        message = {
            'Subject': {
                'Data': subject
            },
            'Body': {
                'Html': {
                    'Data': html_content
                }
            }
        }
        
        # Create raw email message with attachment
        msg_raw = {
            'Source': sender,
            'Destinations': recipients,
            'RawMessage': {
                'Data': create_raw_email_with_attachment(
                    sender=sender,
                    recipients=recipients,
                    subject=subject,
                    html_body=html_content,
                    attachment_data=excel_buffer.getvalue(),
                    attachment_name=excel_filename
                )
            }
        }
        
        # Send email
        response = ses_client.send_raw_email(**msg_raw)
        print(f"Email sent successfully. Message ID: {response['MessageId']}")

        
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        traceback.print_exc()
    
    try:
        # Determine which bucket to use - external or internal
        bucket_name = S3_BUCKET_NAME if S3_BUCKET_NAME else REPORTS_BUCKET
        
        if not bucket_name:
            return False, "No S3 bucket configured (neither S3_BUCKET_NAME nor REPORTS_BUCKET)"
            
        # Create S3 client
        s3_client = boto3.client('s3')
        
        # Generate S3 key with prefix if provided (only for external bucket)
        if S3_BUCKET_NAME:
            s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{excel_filename}" if S3_KEY_PREFIX else excel_filename
        else:
            s3_key = excel_filename  # No prefix for internal bucket
        
        # Reset buffer position to the beginning
        excel_buffer.seek(0)
        
        # Upload buffer to S3 using put_object
        s3_client.put_object(
            Bucket=bucket_name,
            Key=s3_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate S3 URL for the uploaded file
        s3_url = f"s3://{bucket_name}/{s3_key}"
        
        print(f"Successfully uploaded file to {s3_url}")
             
    except Exception as e:
        error_message = f"Error uploading file to S3: {str(e)}"
        print(error_message)
        return False, error_message

def create_raw_email_with_attachment(sender, recipients, subject, html_body, attachment_data, attachment_name):
    """
    Create raw email with attachment
    
    Args:
        sender (str): Sender email
        recipients (list): List of recipient emails
        subject (str): Email subject
        html_body (str): HTML email body
        attachment_data (bytes): Attachment data
        attachment_name (str): Attachment filename
        
    Returns:
        bytes: Raw email message
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    
    # Create message container
    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ', '.join(recipients)
    
    # Create HTML part
    msg_body = MIMEMultipart('alternative')
    html_part = MIMEText(html_body, 'html')
    msg_body.attach(html_part)
    msg.attach(msg_body)
    
    # Create attachment part
    att = MIMEApplication(attachment_data)
    att.add_header('Content-Disposition', 'attachment', filename=attachment_name)
    msg.attach(att)
    
    # Convert to string and return
    return msg.as_string().encode('utf-8')

def add_cloudwatch_metrics(event_categories, analyzed_count, total_count, filtered_count):
    """
    Add CloudWatch metrics for monitoring
    
    Args:
        event_categories (dict): Event categories count
        analyzed_count (int): Number of analyzed events
        total_count (int): Total number of events
        filtered_count (int): Number of filtered events
        
    Returns:
        None
    """
    try:
        # Create CloudWatch client
        cloudwatch = boto3.client('cloudwatch')
        
        # Create metrics data
        metrics_data = [
            {
                'MetricName': 'AnalyzedEvents',
                'Value': analyzed_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'TotalEvents',
                'Value': total_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'FilteredEvents',
                'Value': filtered_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'CriticalEvents',
                'Value': event_categories.get('critical', 0),
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'HighRiskEvents',
                'Value': event_categories.get('high_risk', 0),
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            }
        ]
        
        # Put metrics data
        cloudwatch.put_metric_data(
            Namespace='AWS/HealthEventsAnalysis',
            MetricData=metrics_data
        )
        
        print("CloudWatch metrics published successfully")
        
    except Exception as e:
        print(f"Error publishing CloudWatch metrics: {str(e)}")
        # Don't raise the exception - metrics are non-critical

def fetch_health_event_details1(event_arn, account_id=None):
    """
    Fetch detailed event information from AWS Health API for any account in the organization
    
    Args:
        event_arn (str): ARN of the health event
        account_id (str, optional): AWS account ID that owns the event
        
    Returns:
        dict: Event details including affected resources
    """
    try:
        # Health API should go to us-east-1
        health_client = boto3.client('health', region_name='us-east-1')
        
        # First try organization API (works for both current and linked accounts)
        try:
            # Prepare request for organization event details
            org_filter = {
                'eventArn': event_arn
            }
            
            # Add account ID if provided
            if account_id:
                org_filter['awsAccountId'] = account_id
            
            # Get event details using organization API
            org_event_details = health_client.describe_event_details_for_organization(
                organizationEventDetailFilters=[org_filter]
            )
            
            # Get affected entities using organization API
            org_affected_entities = health_client.describe_affected_entities_for_organization(
                organizationEntityFilters=[
                    {
                        'eventArn': event_arn,
                        'awsAccountId': account_id if account_id else get_account_id_from_event(event_arn)
                    }
                ]
            )
            
            # Check if we got successful results
            if org_event_details.get('successfulSet') and len(org_event_details['successfulSet']) > 0:
                return {
                    'details': org_event_details['successfulSet'][0],
                    'entities': org_affected_entities.get('entities', [])
                }
            
            # If we got here, organization API didn't return results
            print(f"Organization API didn't return results for event {event_arn}")
            
        except Exception as org_error:
            print(f"Error using organization API for event {event_arn}: {str(org_error)}")
        
        # Fall back to account-specific API (only works for current account)
        print(f"Falling back to account-specific API for event {event_arn}")
        
        event_details = health_client.describe_event_details(
            eventArns=[event_arn]
        )
        
        affected_entities = health_client.describe_affected_entities(
            filter={
                'eventArns': [event_arn]
            }
        )
        
        return {
            'details': event_details.get('successfulSet', [{}])[0] if event_details.get('successfulSet') else {},
            'entities': affected_entities.get('entities', [])
        }
        
    except Exception as e:
        print(f"Error fetching Health API data: {str(e)}")
        return {'details': {}, 'entities': []}

def upload_file_to_s3(file_path, file_name):
    """
    Upload a file to S3 bucket using environment variables for configuration
    
    Args:
        file_path: Local path to the file
        file_name: Name to use for the file in S3
        
    Returns:
        tuple: (success boolean, S3 URL or error message)
    """
    try:
        # Determine which bucket to use - external or internal
        bucket_name = S3_BUCKET_NAME if S3_BUCKET_NAME else REPORTS_BUCKET
        
        if not bucket_name:
            return False, "No S3 bucket configured (neither S3_BUCKET_NAME nor REPORTS_BUCKET)"
            
        # Create S3 client
        s3_client = boto3.client('s3')
        
        # Generate S3 key with prefix if provided (only for external bucket)
        if S3_BUCKET_NAME:
            s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{file_name}" if S3_KEY_PREFIX else file_name
        else:
            s3_key = file_name  # No prefix for internal bucket
        
        # Upload file to S3
        s3_client.upload_file(file_path, bucket_name, s3_key)
        
        # Generate S3 URL for the uploaded file
        s3_url = f"s3://{bucket_name}/{s3_key}"
        
        print(f"Successfully uploaded file to {s3_url}")
        return True, s3_url
        
    except Exception as e:
        error_message = f"Error uploading file to S3: {str(e)}"
        print(error_message)
        return False, error_message
