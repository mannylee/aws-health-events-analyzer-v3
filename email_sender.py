"""
Email Sender Lambda Function
Sends individual emails from SQS queue with Excel reports
"""
import boto3
import json
import os
import traceback
from datetime import datetime
from io import BytesIO
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import base64

# Import Excel generation functions from the original code
# We'll need to extract these functions or recreate them
from shared_utils import get_workflow_results


def get_message_retry_count(record):
    """
    Get the retry count for an SQS message based on receive count
    """
    try:
        # SQS provides approximate receive count in message attributes
        attributes = record.get('attributes', {})
        receive_count = int(attributes.get('ApproximateReceiveCount', '1'))
        return receive_count - 1  # Convert to retry count (0-based)
    except (ValueError, KeyError):
        return 0


def lambda_handler(event, context):
    """
    Send individual emails from SQS queue
    """
    successful_messages = []
    failed_messages = []
    
    print(f"Email sender invoked with {len(event['Records'])} SQS messages to process")
    
    for i, record in enumerate(event['Records'], 1):
        message = None
        try:
            message = json.loads(record['body'])
            email_type = message['email_type']
            
            print(f"Processing message {i}/{len(event['Records'])}: {email_type} email")
            
            # Get retry count to implement max retry logic
            retry_count = get_message_retry_count(record)
            max_retries = 3  # Maximum number of retries before giving up
            
            if retry_count >= max_retries:
                print(f"Message for {email_type} email has exceeded {max_retries} retries. Marking as processed to prevent infinite retry.")
                successful_messages.append(record['receiptHandle'])
                continue
            
            if email_type == 'account_specific':
                success = send_account_specific_email(message)
            elif email_type == 'master_report':
                success = send_master_report_email(message)
            else:
                print(f"Unknown email type: {email_type}")
                success = False
            
            if success:
                successful_messages.append(record['receiptHandle'])
                print(f"Successfully sent {email_type} email")
            else:
                # Only retry if we haven't exceeded max attempts
                if retry_count < max_retries - 1:
                    failed_messages.append({
                        'itemIdentifier': record['receiptHandle']
                    })
                    print(f"Failed to send {email_type} email (attempt {retry_count + 1}/{max_retries})")
                else:
                    # Max retries reached - mark as processed to prevent infinite retry
                    successful_messages.append(record['receiptHandle'])
                    print(f"Failed to send {email_type} email after {max_retries} attempts. Marking as processed to prevent infinite retry.")
            
        except Exception as e:
            print(f"Error processing email message: {str(e)}")
            traceback.print_exc()
            
            # Get retry count for exception handling
            retry_count = get_message_retry_count(record) if record else 0
            max_retries = 3
            
            if retry_count >= max_retries - 1:
                # Max retries reached - mark as processed to prevent infinite retry
                successful_messages.append(record['receiptHandle'])
                print(f"Marked message as processed after {retry_count + 1} failed attempts to prevent infinite retry")
            else:
                failed_messages.append({
                    'itemIdentifier': record['receiptHandle']
                })
                print(f"Will retry message (attempt {retry_count + 1}/{max_retries})")
    
    # Return partial batch failure info (SQS will retry failed messages)
    if failed_messages:
        return {
            'batchItemFailures': failed_messages
        }
    
    return {
        'status': 'success',
        'processed': len(successful_messages),
        'failed': len(failed_messages)
    }


def send_account_specific_email(message):
    """
    Send account-specific email
    """
    try:
        # Check if TEST_MASTER_EMAIL_ONLY is enabled
        test_master_only = os.environ.get('TEST_MASTER_EMAIL_ONLY', 'false').lower() == 'true'
        
        if test_master_only:
            print("TEST_MASTER_EMAIL_ONLY enabled - skipping account-specific email")
            return True  # Return success to avoid retries
        
        email_address = message['email_address']
        account_results = message['account_results']
        workflow_id = message['workflow_id']
        
        # Check if this email was already sent (deduplication)
        email_key = f"{workflow_id}-{email_address}-account"
        if check_email_already_sent(email_key):
            print(f"DUPLICATE DETECTED: Account-specific email for {email_address} already sent for workflow {workflow_id}")
            print(f"Skipping duplicate email to prevent spam. Email key: {email_key}")
            return True  # Return success to avoid retries
        
        print(f"Sending account-specific email to {email_address} for {len(account_results)} accounts")
        
        # Extract events and analysis from results
        all_events = []
        all_analyses = []
        account_ids = []
        
        for result in account_results:
            result_data = result.get('result_data', {})
            processed_events = result_data.get('processed_events', [])
            analysis = result_data.get('analysis', {})
            account_id = result.get('account_id', 'unknown')
            
            # Merge Bedrock analysis results back into events for spreadsheet columns
            for event in processed_events:
                if analysis:
                    # Update the existing analysis_summary with Bedrock results
                    current_summary = event.get('analysis_summary', {})
                    
                    # Merge Bedrock analysis fields into analysis_summary
                    current_summary.update({
                        'risk_level': analysis.get('risk_level', current_summary.get('risk_level', 'Low')),
                        'risk_category': analysis.get('risk_category', current_summary.get('risk_category', 'Operational')),
                        'time_sensitivity': analysis.get('time_sensitivity', current_summary.get('time_sensitivity', 'Routine')),
                        'impact_analysis': analysis.get('impact_analysis', ''),
                        'consequences_if_ignored': analysis.get('consequences_if_ignored', ''),
                        'recommended_actions': analysis.get('recommended_actions', []),
                        'manual_review_required': analysis.get('manual_review_required', False)
                    })
                    
                    # Store the full analysis for reference
                    event['analysis'] = analysis
                    event['analysis_summary'] = current_summary
            
            all_events.extend(processed_events)
            all_analyses.append({
                'account_id': account_id,
                'analysis': analysis,
                'event_count': len(processed_events)
            })
            account_ids.append(account_id)
        
        # Generate Excel report for these accounts only
        try:
            excel_buffer = create_account_specific_excel_report(all_events, all_analyses, account_ids)
        except Exception as excel_error:
            print(f"Error creating Excel report for account-specific email: {excel_error}")
            # Create a minimal Excel buffer as fallback
            excel_buffer = create_minimal_excel_fallback(all_events, account_ids)
        
        # Generate HTML content
        try:
            html_content = create_account_specific_html_content(all_analyses, account_ids)
        except Exception as html_error:
            print(f"Error creating HTML content for account-specific email: {html_error}")
            # Create minimal HTML as fallback
            html_content = create_minimal_html_fallback(account_ids, "account-specific")
        
        # Get CC email if configured
        cc_email = os.environ.get('ACCOUNT_SPECIFIC_EMAIL_CC', '').strip()
        cc_recipients = [cc_email] if cc_email else []
        
        if cc_recipients:
            print(f"Account-specific email will be CC'd to: {cc_email}")
        else:
            print("No CC configured for account-specific emails")
        
        # Send email
        success = send_email_with_attachment(
            recipients=[email_address],
            cc_recipients=cc_recipients,
            subject=create_email_subject(all_events, all_analyses, account_ids),
            html_content=html_content,
            excel_buffer=excel_buffer,
            filename_suffix=f"_accounts_{'_'.join(account_ids[:3])}"  # Limit filename length
        )
        
        # Mark email as sent if successful
        if success:
            mark_email_as_sent(email_key)
        
        return success
        
    except Exception as e:
        print(f"Error sending account-specific email: {str(e)}")
        traceback.print_exc()
        return False


def send_master_report_email(message):
    """
    Send master report email with all results
    """
    try:
        email_addresses = message['email_addresses']
        all_results = message['all_results']
        workflow_id = message['workflow_id']
        
        # Check if this master report was already sent (deduplication)
        email_key = f"{workflow_id}-master-report"
        if check_email_already_sent(email_key):
            print(f"DUPLICATE DETECTED: Master report email already sent for workflow {workflow_id}")
            print(f"Skipping duplicate master report to prevent spam. Email key: {email_key}")
            return True  # Return success to avoid retries
        
        print(f"Sending master report to {len(email_addresses)} recipients for {len(all_results)} accounts")
        
        # Extract all events and analyses
        all_events = []
        all_analyses = []
        account_ids = []
        
        for result in all_results:
            result_data = result.get('result_data', {})
            processed_events = result_data.get('processed_events', [])
            analysis = result_data.get('analysis', {})
            account_id = result.get('account_id', 'unknown')
            email_mapping = result_data.get('email_mapping')
            
            # Merge Bedrock analysis results back into events for spreadsheet columns
            for event in processed_events:
                # Add tracking information
                event['mapped_email'] = email_mapping or 'Default Recipients'
                
                # Determine email status based on SES mode and verification
                email_status = determine_email_status(email_mapping)
                event['email_sent_status'] = email_status
                
                # Merge Bedrock analysis results into analysis_summary for spreadsheet
                if analysis:
                    # Update the existing analysis_summary with Bedrock results
                    current_summary = event.get('analysis_summary', {})
                    
                    # Merge Bedrock analysis fields into analysis_summary
                    current_summary.update({
                        'risk_level': analysis.get('risk_level', current_summary.get('risk_level', 'Low')),
                        'risk_category': analysis.get('risk_category', current_summary.get('risk_category', 'Operational')),
                        'time_sensitivity': analysis.get('time_sensitivity', current_summary.get('time_sensitivity', 'Routine')),
                        'impact_analysis': analysis.get('impact_analysis', ''),
                        'consequences_if_ignored': analysis.get('consequences_if_ignored', ''),
                        'recommended_actions': analysis.get('recommended_actions', []),
                        'manual_review_required': analysis.get('manual_review_required', False)
                    })
                    
                    # Store the full analysis for reference
                    event['analysis'] = analysis
                    event['analysis_summary'] = current_summary
            
            all_events.extend(processed_events)
            all_analyses.append({
                'account_id': account_id,
                'analysis': analysis,
                'event_count': len(processed_events),
                'email_mapping': email_mapping
            })
            account_ids.append(account_id)
        
        # Generate comprehensive Excel report with tracking
        try:
            excel_buffer = create_master_excel_report(all_events, all_analyses, account_ids, all_results)
        except Exception as excel_error:
            print(f"Error creating Excel report for master email: {excel_error}")
            # Create a minimal Excel buffer as fallback
            excel_buffer = create_minimal_excel_fallback(all_events, account_ids)
        
        # Generate HTML content
        try:
            html_content = create_master_html_content(all_analyses, account_ids)
        except Exception as html_error:
            print(f"Error creating HTML content for master email: {html_error}")
            # Create minimal HTML as fallback
            html_content = create_minimal_html_fallback(account_ids, "master report")
        
        # Send email
        success = send_email_with_attachment(
            recipients=email_addresses,
            subject=create_email_subject(all_events, all_analyses, account_ids, is_master=True),
            html_content=html_content,
            excel_buffer=excel_buffer,
            filename_suffix="_master_report"
        )
        
        # Mark email as sent if successful
        if success:
            mark_email_as_sent(email_key)
        
        return success
        
    except Exception as e:
        print(f"Error sending master report email: {str(e)}")
        traceback.print_exc()
        return False


def create_account_specific_excel_report(events, analyses, account_ids):
    """
    Create Excel report for specific accounts (simplified version)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary sheet
    summary_sheet = wb.create_sheet("Summary")
    create_summary_sheet(summary_sheet, events, analyses, account_ids, is_master=False)
    
    # Events sheet
    events_sheet = wb.create_sheet("Events")
    create_events_sheet(events_sheet, events, include_tracking=False)
    
    # Analysis sheet
    analysis_sheet = wb.create_sheet("Risk Analysis")
    create_analysis_sheet(analysis_sheet, analyses)
    
    # Save to buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_master_excel_report(events, analyses, account_ids, all_results):
    """
    Create comprehensive Excel report with tracking (master version)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary sheet with configuration details
    summary_sheet = wb.create_sheet("Summary")
    create_summary_sheet(summary_sheet, events, analyses, account_ids, is_master=True)
    
    # All Events sheet with tracking columns
    events_sheet = wb.create_sheet("All Events")
    create_events_sheet(events_sheet, events, include_tracking=True)
    
    # Critical Events sheet (always create, even if empty)
    critical_events = [e for e in events if e.get('analysis_summary', {}).get('severity') == 'critical' or 
                      e.get('analysis_summary', {}).get('risk_level') == 'Critical']
    critical_sheet = wb.create_sheet("Critical Events")
    create_events_sheet(critical_sheet, critical_events, include_tracking=True)
    
    # Risk Analysis sheet
    analysis_sheet = wb.create_sheet("Risk Analysis")
    create_analysis_sheet(analysis_sheet, analyses)
    
    # Account Email Mapping sheet (last sheet as requested)
    mapping_sheet = wb.create_sheet("Account Email Mapping")
    create_account_mapping_sheet(mapping_sheet, all_results)
    
    # Save to buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_summary_sheet(sheet, events, analyses, account_ids, is_master=False):
    """
    Create summary sheet
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Title
    sheet['A1'] = "AWS Health Events Analysis Summary"
    sheet['A1'].font = Font(size=16, bold=True)
    
    row = 3
    
    # Check for manual review requirements
    manual_review_count = len([a for a in analyses if a.get('analysis', {}).get('manual_review_required', False)])
    
    if manual_review_count > 0:
        sheet[f'A{row}'] = "⚠️ ATTENTION: Manual Review Required"
        sheet[f'A{row}'].font = Font(size=14, bold=True, color="FF6B35")
        sheet[f'A{row}'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        row += 1
        sheet[f'A{row}'] = f"{manual_review_count} account(s) require manual review due to AI analysis unavailability"
        sheet[f'A{row}'].font = Font(color="856404")
        row += 2
    
    # Basic statistics
    sheet[f'A{row}'] = "Analysis Statistics"
    sheet[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    sheet[f'A{row}'] = f"Total Events: {len(events)}"
    row += 1
    sheet[f'A{row}'] = f"Total Accounts: {len(account_ids)}"
    row += 1
    
    if manual_review_count > 0:
        sheet[f'A{row}'] = f"Manual Review Required: {manual_review_count} accounts"
        sheet[f'A{row}'].font = Font(color="FF6B35")
        row += 1
    
    # Event severity breakdown
    severity_counts = {}
    for event in events:
        severity = event.get('severity', 'low')
        severity_counts[severity] = severity_counts.get(severity, 0) + 1
    
    for severity, count in severity_counts.items():
        sheet[f'A{row}'] = f"{severity.title()} Events: {count}"
        row += 1
    
    row += 1
    
    # Risk analysis summary
    sheet[f'A{row}'] = "Risk Analysis Summary"
    sheet[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    risk_counts = {}
    for analysis in analyses:
        analysis_data = analysis.get('analysis', {})
        if not analysis_data.get('manual_review_required', False):
            risk_level = analysis_data.get('risk_level', 'Low')
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1
    
    for risk, count in risk_counts.items():
        sheet[f'A{row}'] = f"{risk} Risk Accounts: {count}"
        row += 1
    
    # Configuration info for master reports
    if is_master:
        row += 5  # Move Configuration section down to row 18 (was row 15)
        sheet[f'A{row}'] = "Configuration"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        sheet[f'A{row}'] = f"Custom Mapping Enabled: {os.environ.get('USE_CUSTOM_ACCOUNT_EMAIL_MAPPING', 'false')}"
        row += 1
        sheet[f'A{row}'] = f"Organizations Mapping Enabled: {os.environ.get('USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING', 'false')}"
        row += 1
        sheet[f'A{row}'] = f"Default Recipients: {os.environ.get('RECIPIENT_EMAILS', 'Not configured')}"
        row += 1


def create_events_sheet(sheet, events, include_tracking=False):
    """
    Create events sheet with correct column structure
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Headers as specified
    headers = [
        "Event ARN", "Event Type", "Region", "Start Time", "Last Update", 
        "Category", "Description", "Critical", "Risk Level", "Account ID",
        "Key Date", "Time Sensitivity", "Risk Category", "Required Actions",
        "Impact Analysis", "Consequences If Ignored", "Affected Resources"
    ]
    
    if include_tracking:
        headers.extend(["Mapped Email", "Email Status"])
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data
    for row, event in enumerate(events, 2):
        analysis = event.get('analysis_summary', {})
        
        sheet.cell(row=row, column=1, value=event.get('arn', ''))
        sheet.cell(row=row, column=2, value=event.get('eventTypeCode', ''))
        sheet.cell(row=row, column=3, value=event.get('region', ''))
        sheet.cell(row=row, column=4, value=str(event.get('startTime', '')))
        sheet.cell(row=row, column=5, value=str(event.get('lastUpdatedTime', '')))
        sheet.cell(row=row, column=6, value=event.get('eventTypeCategory', ''))
        
        # Try multiple sources for event description
        description = (
            event.get('eventDescription', '') or 
            event.get('description', '') or 
            analysis.get('description_preview', '') or
            analysis.get('description', '') or
            event.get('message', '') or
            ''
        )
        sheet.cell(row=row, column=7, value=description[:500] if description else 'No description available')
        
        # Critical flag
        is_critical = analysis.get('severity') == 'critical' or analysis.get('risk_level') == 'Critical'
        sheet.cell(row=row, column=8, value='Yes' if is_critical else 'No')
        
        # Risk level with manual review indicator
        risk_level = analysis.get('risk_level', 'Low')
        manual_review = event.get('analysis', {}).get('manual_review_required', False)
        if manual_review:
            risk_level = f"MANUAL REVIEW - {risk_level}"
        sheet.cell(row=row, column=9, value=risk_level)
        sheet.cell(row=row, column=10, value=event.get('accountId', ''))
        sheet.cell(row=row, column=11, value=str(event.get('startTime', '')))  # Key Date same as Start Time
        sheet.cell(row=row, column=12, value=analysis.get('time_sensitivity', 'Normal'))
        sheet.cell(row=row, column=13, value=analysis.get('risk_category', 'Operational'))
        
        # Required Actions
        actions = analysis.get('recommended_actions', [])
        if isinstance(actions, list) and actions:
            # Handle list of action objects
            actions_text = '; '.join([f"{a.get('action', '') if isinstance(a, dict) else str(a)}" for a in actions[:3] if a])
        elif isinstance(actions, str):
            # Handle string directly
            actions_text = actions[:200]
        else:
            # Fallback to required_actions field if recommended_actions is not available
            required_actions = analysis.get('required_actions', '')
            actions_text = str(required_actions)[:200] if required_actions else ''
        sheet.cell(row=row, column=14, value=actions_text)
        
        # Impact Analysis - ensure it's a string before slicing
        impact_analysis = analysis.get('impact_analysis', '') or ''
        sheet.cell(row=row, column=15, value=str(impact_analysis)[:300])
        
        # Consequences If Ignored - ensure it's a string before slicing
        consequences = analysis.get('consequences_if_ignored', '') or ''
        sheet.cell(row=row, column=16, value=str(consequences)[:300])
        
        # Affected Resources - show actual resource ARNs/IDs
        affected_entities = event.get('affectedEntities', [])
        if affected_entities:
            # Extract resource ARNs/IDs from affected entities
            resources = []
            for entity in affected_entities[:5]:  # Limit to first 5 resources
                entity_value = entity.get('entityValue', '')
                entity_url = entity.get('entityUrl', '')
                
                # Prefer entityValue (usually contains ARN or resource ID)
                if entity_value:
                    resources.append(entity_value)
                elif entity_url:
                    resources.append(entity_url)
            
            resources_text = ', '.join(resources) if resources else f"Account: {event.get('accountId', 'Unknown')}"
        else:
            # Fallback: try to get from event description or use account ID
            event_description = event.get('eventDescription', '')
            
            # Look for ARN patterns in the description
            import re
            arn_pattern = r'arn:aws:[a-zA-Z0-9\-]+:[a-zA-Z0-9\-]*:\d*:[a-zA-Z0-9\-/._]+'
            arns = re.findall(arn_pattern, event_description)
            
            if arns:
                resources_text = ', '.join(arns[:3])  # Limit to first 3 ARNs found
            else:
                resources_text = f"Account: {event.get('accountId', 'Unknown')}"
        
        sheet.cell(row=row, column=17, value=resources_text)
        
        if include_tracking:
            sheet.cell(row=row, column=18, value=event.get('mapped_email', ''))
            sheet.cell(row=row, column=19, value=event.get('email_sent_status', ''))


def create_analysis_sheet(sheet, analyses):
    """
    Create risk analysis sheet with correct column structure
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Headers as specified
    headers = ["Account ID", "Event Type", "Region", "Risk Level", "Analysis Status", "Full Analysis"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data - one row per account with analysis summary
    row = 2
    for analysis in analyses:
        analysis_data = analysis.get('analysis', {})
        account_id = analysis.get('account_id', '')
        manual_review = analysis_data.get('manual_review_required', False)
        
        # Account ID
        sheet.cell(row=row, column=1, value=account_id)
        
        # Event Type - summarize
        processed_events = analysis_data.get('processed_events', [])
        if processed_events:
            event_types = list(set([e.get('eventTypeCode', '') for e in processed_events[:3]]))
            event_type_summary = ', '.join(event_types)
            if len(processed_events) > 3:
                event_type_summary += f" (+{len(processed_events)-3} more)"
        else:
            event_type_summary = 'Multiple Events'
        sheet.cell(row=row, column=2, value=event_type_summary)
        
        # Region - summarize
        if processed_events:
            regions = list(set([e.get('region', '') for e in processed_events[:3]]))
            region_summary = ', '.join(regions)
            if len(set([e.get('region', '') for e in processed_events])) > 3:
                region_summary += " (+more)"
        else:
            region_summary = 'Multiple Regions'
        sheet.cell(row=row, column=3, value=region_summary)
        
        # Risk Level
        risk_level = analysis_data.get('risk_level', 'Low')
        if manual_review:
            risk_level = f"MANUAL REVIEW - {risk_level}"
        sheet.cell(row=row, column=4, value=risk_level)
        
        # Analysis Status
        if manual_review:
            status = "⚠️ PLACEHOLDER - Manual Review Required"
            cell = sheet.cell(row=row, column=5, value=status)
            cell.font = Font(color="FF6B35")
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            status = "✅ AI Analysis Complete"
            cell = sheet.cell(row=row, column=5, value=status)
            cell.font = Font(color="28A745")
        
        # Full Analysis - JSON response
        if manual_review:
            # For placeholder analysis, include the reason and guidance
            full_analysis = {
                'account_id': account_id,
                'analysis_status': 'PLACEHOLDER_ANALYSIS',
                'manual_review_required': True,
                'placeholder_reason': analysis_data.get('placeholder_reason', 'AI analysis unavailable'),
                'failure_reason': analysis_data.get('failure_reason', 'Unknown'),
                'guidance': 'Please review all events manually and consult AWS Health Dashboard',
                'analysis_data': analysis_data
            }
        else:
            # Regular analysis
            full_analysis = {
                'account_id': account_id,
                'analysis_status': 'AI_ANALYSIS_COMPLETE',
                'analysis_data': analysis_data
            }
        
        full_analysis_json = json.dumps(full_analysis, indent=2, default=str)
        sheet.cell(row=row, column=6, value=full_analysis_json)
        row += 1


def create_account_mapping_sheet(sheet, all_results):
    """
    Create account email mapping sheet showing all available mappings from both sources
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from shared_utils import get_combined_account_email_mapping
    
    # Headers
    headers = ["Account ID", "Email Address", "Mapping Source"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Get all available mappings from both sources
    try:
        account_to_email, email_to_accounts, mapping_sources = get_combined_account_email_mapping()
        
        # Create comprehensive mapping list
        mappings = []
        
        # Add all accounts that have mappings
        for account_id, email in account_to_email.items():
            source = mapping_sources.get(account_id, 'unknown')
            
            # Convert source codes to readable descriptions
            if source == 'custom':
                source_desc = 'DynamoDB Only'
            elif source == 'organizations':
                source_desc = 'AWS Organizations Only'
            elif source == 'combined':
                source_desc = 'Both (DynamoDB takes precedence)'
            else:
                source_desc = 'Unknown Source'
            
            mappings.append({
                'account_id': account_id,
                'email': email,
                'source': source_desc
            })
        
        # Add accounts from workflow results that don't have mappings
        workflow_accounts = set()
        for result in all_results:
            account_id = result.get('account_id', 'unknown')
            workflow_accounts.add(account_id)
            
            # If this account wasn't in the mappings, add it as default
            if account_id not in account_to_email and account_id != 'unknown':
                mappings.append({
                    'account_id': account_id,
                    'email': 'Default Recipients',
                    'source': 'No mapping found'
                })
        
        # Sort by source type, then email, then account
        def sort_key(mapping):
            source_priority = {
                'Both (DynamoDB takes precedence)': 1,
                'DynamoDB Only': 2,
                'AWS Organizations Only': 3,
                'No mapping found': 4,
                'Unknown Source': 5
            }
            return (source_priority.get(mapping['source'], 6), mapping['email'], mapping['account_id'])
        
        mappings.sort(key=sort_key)
        
        # Write data
        for row, mapping in enumerate(mappings, 2):
            sheet.cell(row=row, column=1, value=mapping['account_id'])
            sheet.cell(row=row, column=2, value=mapping['email'])
            
            # Color code the source column
            source_cell = sheet.cell(row=row, column=3, value=mapping['source'])
            
            if mapping['source'] == 'Both (DynamoDB takes precedence)':
                source_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                source_cell.font = Font(color="0066CC")
            elif mapping['source'] == 'DynamoDB Only':
                source_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                source_cell.font = Font(color="006600")
            elif mapping['source'] == 'AWS Organizations Only':
                source_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                source_cell.font = Font(color="FF8C00")
            elif mapping['source'] == 'No mapping found':
                source_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                source_cell.font = Font(color="666666")
        
        # Add configuration info at the bottom
        if mappings:
            config_row = len(mappings) + 4
            
            # Configuration header
            config_header = sheet.cell(row=config_row, column=1, value="Configuration:")
            config_header.font = Font(bold=True, size=12)
            config_row += 1
            
            # Show current configuration
            use_custom = os.environ.get('USE_CUSTOM_ACCOUNT_EMAIL_MAPPING', 'false').lower() == 'true'
            use_org = os.environ.get('USE_ORGANIZATION_ACCOUNT_EMAIL_MAPPING', 'false').lower() == 'true'
            
            sheet.cell(row=config_row, column=1, value=f"DynamoDB Mapping Enabled: {'Yes' if use_custom else 'No'}")
            config_row += 1
            sheet.cell(row=config_row, column=1, value=f"AWS Organizations Mapping Enabled: {'Yes' if use_org else 'No'}")
            config_row += 1
            
            if use_custom and use_org:
                sheet.cell(row=config_row, column=1, value="Priority: DynamoDB overrides AWS Organizations when both exist")
            elif use_custom:
                sheet.cell(row=config_row, column=1, value="Source: DynamoDB mappings only")
            elif use_org:
                sheet.cell(row=config_row, column=1, value="Source: AWS Organizations mappings only")
            else:
                sheet.cell(row=config_row, column=1, value="Source: No automatic mapping configured")
        
    except Exception as e:
        print(f"Error creating account mapping sheet: {e}")
        # Fallback to simple mapping from workflow results
        mappings = []
        for result in all_results:
            result_data = result.get('result_data', {})
            account_id = result.get('account_id', 'unknown')
            email_mapping = result_data.get('email_mapping')
            
            if email_mapping:
                mappings.append({
                    'account_id': account_id,
                    'email': email_mapping,
                    'source': 'From Workflow Results'
                })
            else:
                mappings.append({
                    'account_id': account_id,
                    'email': 'Default Recipients',
                    'source': 'No mapping found'
                })
        
        # Write fallback data
        for row, mapping in enumerate(mappings, 2):
            sheet.cell(row=row, column=1, value=mapping['account_id'])
            sheet.cell(row=row, column=2, value=mapping['email'])
            sheet.cell(row=row, column=3, value=mapping['source'])


def create_account_specific_html_content(analyses, account_ids):
    """
    Create HTML content for account-specific emails
    """
    customer_name = os.environ.get('CUSTOMER_NAME', 'AWS Health Events')
    
    # Check for manual review requirements
    manual_review_accounts = [a for a in analyses if a.get('analysis', {}).get('manual_review_required', False)]
    
    html = f"""
    <html>
    <body>
        <h2>{customer_name} - AWS Health Events Analysis</h2>
        <p>This report contains AWS Health events analysis for your accounts: {', '.join(account_ids[:5])}</p>
    """
    
    if manual_review_accounts:
        html += f"""
        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; margin: 10px 0; border-radius: 5px;">
            <h3 style="color: #856404; margin-top: 0;">⚠️ Manual Review Required</h3>
            <p style="color: #856404; margin-bottom: 0;">
                <strong>{len(manual_review_accounts)} account(s)</strong> require manual review because AI analysis was unavailable. 
                Please review the detailed events in the attached Excel report and consult the AWS Health Dashboard.
            </p>
        </div>
        """
    
    html += """
        <h3>Summary</h3>
        <ul>
    """
    
    for analysis in analyses:
        account_id = analysis['account_id']
        analysis_data = analysis.get('analysis', {})
        risk_level = analysis_data.get('risk_level', 'Low')
        event_count = analysis.get('event_count', 0)
        manual_review = analysis_data.get('manual_review_required', False)
        
        status_indicator = "🔍 MANUAL REVIEW" if manual_review else f"{risk_level} risk"
        html += f"<li><strong>Account {account_id}</strong>: {status_indicator}, {event_count} events</li>"
    
    html += """
        </ul>
        
        <p>Please see the attached Excel report for detailed analysis and recommended actions.</p>
        
        <p>This is an automated report from the AWS Health Events Analyzer.</p>
    </body>
    </html>
    """
    
    return html


def create_master_html_content(analyses, account_ids):
    """
    Create HTML content for master report emails
    """
    customer_name = os.environ.get('CUSTOMER_NAME', 'AWS Health Events')
    
    # Calculate summary statistics
    total_accounts = len(account_ids)
    total_events = sum(a.get('event_count', 0) for a in analyses)
    
    risk_counts = {}
    manual_review_count = 0
    
    for analysis in analyses:
        analysis_data = analysis.get('analysis', {})
        if analysis_data.get('manual_review_required', False):
            manual_review_count += 1
        else:
            risk_level = analysis_data.get('risk_level', 'Low')
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1
    
    html = f"""
    <html>
    <body>
        <h2>{customer_name} - AWS Health Events Master Report</h2>
        <p>This is the comprehensive master report containing all AWS Health events across your organization.</p>
    """
    
    if manual_review_count > 0:
        html += f"""
        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; margin: 15px 0; border-radius: 5px;">
            <h3 style="color: #856404; margin-top: 0;">⚠️ Manual Review Required</h3>
            <p style="color: #856404; margin-bottom: 0;">
                <strong>{manual_review_count} account(s)</strong> require manual review because AI analysis was unavailable. 
                These accounts are marked in the Excel report and should be prioritized for manual assessment.
            </p>
        </div>
        """
    
    html += f"""
        <h3>Executive Summary</h3>
        <ul>
            <li><strong>Total Accounts Analyzed:</strong> {total_accounts}</li>
            <li><strong>Total Events:</strong> {total_events}</li>
        </ul>
        
        <h3>Analysis Status</h3>
        <ul>
    """
    
    if manual_review_count > 0:
        html += f"<li><strong>Manual Review Required:</strong> {manual_review_count} accounts</li>"
    
    html += """
        </ul>
        
        <h3>Risk Distribution</h3>
        <ul>
    """
    
    for risk_level in ['Critical', 'High', 'Medium', 'Low']:
        count = risk_counts.get(risk_level, 0)
        if count > 0:
            html += f"<li><strong>{risk_level} Risk Accounts:</strong> {count}</li>"
    
    html += """
        </ul>
        
        <h3>Report Contents</h3>
        <p>The attached Excel report includes:</p>
        <ul>
            <li><strong>Summary Sheet:</strong> Configuration and statistics overview</li>
            <li><strong>All Events Sheet:</strong> Complete event listing with email routing tracking</li>
            <li><strong>Critical Events Sheet:</strong> High-priority events requiring immediate attention</li>
            <li><strong>Account Email Mapping Sheet:</strong> Account-to-email routing configuration</li>
            <li><strong>Risk Analysis Sheet:</strong> AI-powered risk assessment and recommendations</li>
        </ul>
        
        <p>This is an automated report from the AWS Health Events Analyzer.</p>
    </body>
    </html>
    """
    
    return html


def create_email_subject(events, analyses, account_ids, is_master=False):
    """
    Create email subject line
    """
    customer_name = os.environ.get('CUSTOMER_NAME', 'AWS Health Events')
    
    # Count events by severity and risk
    critical_events = len([e for e in events if e.get('severity') == 'critical'])
    high_events = len([e for e in events if e.get('severity') == 'high'])
    total_events = len(events)
    
    if is_master:
        prefix = f"{customer_name} [MASTER]"
    else:
        prefix = customer_name
    
    if critical_events > 0:
        return f"{prefix} [CRITICAL] AWS Health Events - {critical_events} Critical Events"
    elif high_events > 0:
        return f"{prefix} [HIGH RISK] AWS Health Events - {high_events} High Priority Events"
    else:
        return f"{prefix} AWS Health Events Analysis - {total_events} Events"


def check_email_already_sent(email_key):
    """
    Check if an email was already sent using DynamoDB
    """
    try:
        dynamodb = boto3.resource('dynamodb')
        table_name = os.environ.get('WORKFLOW_RESULTS_TABLE')
        table = dynamodb.Table(table_name)
        
        # Check for email sent marker
        response = table.get_item(
            Key={
                'workflow_id': email_key,
                'account_id': 'EMAIL_SENT_MARKER'
            }
        )
        
        return 'Item' in response
        
    except Exception as e:
        print(f"Error checking email sent status for {email_key}: {e}")
        return False  # Default to not sent if check fails


def mark_email_as_sent(email_key):
    """
    Mark an email as sent using DynamoDB with conditional write to prevent race conditions
    """
    try:
        import time
        from botocore.exceptions import ClientError
        dynamodb = boto3.resource('dynamodb')
        table_name = os.environ.get('WORKFLOW_RESULTS_TABLE')
        table = dynamodb.Table(table_name)
        
        # Use conditional write to prevent race conditions
        try:
            table.put_item(
                Item={
                    'workflow_id': email_key,
                    'account_id': 'EMAIL_SENT_MARKER',
                    'result_data': {
                        'email_sent_at': datetime.now().isoformat(),
                        'status': 'email_sent'
                    },
                    'status': 'EMAIL_SENT',
                    'timestamp': datetime.now().isoformat(),
                    'ttl': int(time.time()) + (7 * 24 * 60 * 60)  # 7 days TTL
                },
                ConditionExpression='attribute_not_exists(workflow_id)'
            )
            print(f"Marked email as sent: {email_key}")
            return True
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'ConditionalCheckFailedException':
                print(f"Email already marked as sent for {email_key} (race condition avoided)")
                return False
            else:
                raise
        
    except Exception as e:
        print(f"Error marking email as sent for {email_key}: {e}")
        return False


def create_minimal_excel_fallback(events, account_ids):
    """
    Create a minimal Excel report when full report generation fails
    """
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Health Events Summary"
        
        # Simple headers
        ws['A1'] = "Account ID"
        ws['B1'] = "Event Count"
        ws['C1'] = "Status"
        
        # Count events per account
        account_event_counts = {}
        for event in events:
            account_id = event.get('accountId', 'unknown')
            account_event_counts[account_id] = account_event_counts.get(account_id, 0) + 1
        
        # Write data
        row = 2
        for account_id in account_ids:
            ws[f'A{row}'] = account_id
            ws[f'B{row}'] = account_event_counts.get(account_id, 0)
            ws[f'C{row}'] = "Report generation failed - manual review required"
            row += 1
        
        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        print(f"Error creating minimal Excel fallback: {e}")
        # Return empty buffer as last resort
        return BytesIO()


def create_minimal_html_fallback(account_ids, email_type):
    """
    Create minimal HTML content when full HTML generation fails
    """
    try:
        html_content = f"""
        <html>
        <body>
        <h2>AWS Health Events Analysis - {email_type.title()}</h2>
        <p><strong>⚠️ Report Generation Issue</strong></p>
        <p>There was an issue generating the full report content. Please find the basic information below:</p>
        
        <h3>Accounts Processed:</h3>
        <ul>
        """
        
        for account_id in account_ids[:10]:  # Limit to first 10 accounts
            html_content += f"<li>{account_id}</li>"
        
        if len(account_ids) > 10:
            html_content += f"<li>... and {len(account_ids) - 10} more accounts</li>"
        
        html_content += """
        </ul>
        
        <p><strong>Action Required:</strong> Please check the attached Excel file for event details, or contact your AWS administrator for assistance.</p>
        
        <p>This email was generated automatically by the AWS Health Events Analyzer.</p>
        </body>
        </html>
        """
        
        return html_content
        
    except Exception as e:
        print(f"Error creating minimal HTML fallback: {e}")
        return "<html><body><h2>AWS Health Events Analysis</h2><p>Report generation failed. Please contact your administrator.</p></body></html>"


def determine_email_status(email_mapping):
    """
    Determine email status based on SES mode and email verification
    """
    try:
        # Check if TEST_MASTER_EMAIL_ONLY is enabled
        test_master_only = os.environ.get('TEST_MASTER_EMAIL_ONLY', 'false').lower() == 'true'
        
        if test_master_only and email_mapping:
            # Account-specific emails are skipped in test mode
            return 'SKIPPED - TEST_MASTER_EMAIL_ONLY enabled'
        
        # Check if SES is in sandbox mode
        if not should_verify_recipient_email():
            # Production mode - emails will be sent
            if email_mapping:
                return 'Sent to mapped email'
            else:
                return 'Sent to default recipients'
        
        # Sandbox mode - check verification status
        ses_client = boto3.client('ses')
        
        # Get the actual email addresses to check
        if email_mapping:
            emails_to_check = [email_mapping]
            base_status = 'mapped email'
        else:
            # Check default recipients
            default_recipients = os.environ.get('RECIPIENT_EMAILS', '').split(',')
            emails_to_check = [email.strip() for email in default_recipients if email.strip()]
            base_status = 'default recipients'
        
        # Check verification status for all relevant emails
        unverified_emails = []
        for email in emails_to_check:
            if email:
                try:
                    response = ses_client.get_identity_verification_attributes(
                        Identities=[email]
                    )
                    
                    verification_status = response.get('VerificationAttributes', {}).get(
                        email, {}
                    ).get('VerificationStatus', 'NotStarted')
                    
                    if verification_status != 'Success':
                        unverified_emails.append(email)
                        
                except Exception as e:
                    print(f"Could not verify {email}: {e}")
                    unverified_emails.append(email)
        
        # Determine status based on verification results
        if unverified_emails:
            if len(unverified_emails) == len(emails_to_check):
                # All emails unverified
                return f'UNSENT - {base_status} unverified (SES sandbox)'
            else:
                # Some emails unverified
                return f'PARTIAL - some {base_status} unverified (SES sandbox)'
        else:
            # All emails verified
            return f'Sent to {base_status} (verified)'
            
    except Exception as e:
        print(f"Error determining email status: {e}")
        # Fallback to original logic
        if email_mapping:
            return 'Sent to mapped email (status unknown)'
        else:
            return 'Sent to default recipients (status unknown)'


def send_email_with_attachment(recipients, subject, html_content, excel_buffer, filename_suffix="", cc_recipients=None):
    """
    Send email with Excel attachment using SES
    
    Args:
        recipients (list): Primary recipients (TO field)
        subject (str): Email subject
        html_content (str): HTML email body
        excel_buffer (BytesIO): Excel attachment data
        filename_suffix (str): Optional suffix for filename
        cc_recipients (list): Optional CC recipients
    """
    try:
        sender = os.environ['SENDER_EMAIL']
        
        # Generate filename
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H-%M')
        filename_template = os.environ.get('EXCEL_FILENAME_TEMPLATE', 'AWS_Health_Events_Analysis_{date}_{time}.xlsx')
        filename = filename_template.format(date=current_date, time=current_time)
        
        if filename_suffix:
            # Insert suffix before .xlsx
            filename = filename.replace('.xlsx', f'{filename_suffix}.xlsx')
        
        # Create SES client
        ses_client = boto3.client('ses')
        
        # Prepare all recipients for SES (TO + CC combined)
        cc_recipients = cc_recipients or []
        all_recipients = recipients + cc_recipients
        
        # Check SES mode and verify recipients if needed
        if should_verify_recipient_email():
            print("SES is in sandbox mode - checking recipient verification")
            unverified = check_recipient_verification(ses_client, all_recipients)
            if unverified:
                print(f"Warning: Unverified recipients: {unverified}")
                return False
        
        # Create raw email with attachment (CC header is set inside for display)
        raw_message = create_raw_email_with_attachment(
            sender=sender,
            recipients=recipients,
            cc_recipients=cc_recipients,
            subject=subject,
            html_body=html_content,
            attachment_data=excel_buffer.getvalue(),
            attachment_name=filename
        )
        
        # Send email - SES Destinations includes both TO and CC recipients
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=all_recipients,  # This is the actual delivery list for SES
            RawMessage={'Data': raw_message}
        )
        
        print(f"Email sent successfully. MessageId: {response['MessageId']}")
        return True
        
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        traceback.print_exc()
        return False


def should_verify_recipient_email():
    """
    Check if SES is in sandbox mode
    """
    try:
        ses_client = boto3.client('ses')
        quota = ses_client.get_send_quota()
        
        max_24_hour_send = quota.get('Max24HourSend', 0)
        max_send_rate = quota.get('MaxSendRate', 0)
        
        is_sandbox = (max_24_hour_send <= 200 and max_send_rate <= 1)
        return is_sandbox
        
    except Exception as e:
        print(f"Warning: Could not determine SES mode: {e}")
        return True  # Default to sandbox mode for safety


def check_recipient_verification(ses_client, recipients):
    """
    Check which recipients are not verified
    """
    unverified = []
    
    for recipient in recipients:
        try:
            response = ses_client.get_identity_verification_attributes(
                Identities=[recipient]
            )
            
            verification_status = response.get('VerificationAttributes', {}).get(
                recipient, {}
            ).get('VerificationStatus', 'NotStarted')
            
            if verification_status != 'Success':
                unverified.append(f"{recipient} ({verification_status})")
                
        except Exception as e:
            print(f"Could not verify {recipient}: {e}")
            unverified.append(f"{recipient} (check failed)")
    
    return unverified


def create_raw_email_with_attachment(sender, recipients, subject, html_body, attachment_data, attachment_name, cc_recipients=None):
    """
    Create raw email message with attachment
    
    Args:
        sender (str): Sender email address
        recipients (list): Primary recipients (TO field)
        subject (str): Email subject
        html_body (str): HTML email body
        attachment_data (bytes): Excel attachment data
        attachment_name (str): Attachment filename
        cc_recipients (list): Optional CC recipients
    """
    # Create message container
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ', '.join(recipients)
    
    # Add CC header if CC recipients are provided
    cc_recipients = cc_recipients or []
    if cc_recipients:
        msg['Cc'] = ', '.join(cc_recipients)
        print(f"Adding CC recipients: {', '.join(cc_recipients)}")
    
    # Add HTML body
    msg.attach(MIMEText(html_body, 'html'))
    
    # Add Excel attachment
    attachment = MIMEApplication(attachment_data)
    attachment.add_header('Content-Disposition', 'attachment', filename=attachment_name)
    msg.attach(attachment)
    
    return msg.as_string()